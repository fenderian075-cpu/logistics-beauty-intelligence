#!/usr/bin/env python3
"""Parse the official e-Stat IIP 2020-base workbook (2018-present).

Workbook contract observed from e-Stat:
- sheets: 生産 / 出荷 / 在庫 / 在庫率
- row 3: monthly period headers as YYYYMM from column D onward
- row 4: 鉱工業 aggregate
No values are inferred.
"""
from __future__ import annotations
import io, json, re, requests
from openpyxl import load_workbook
from backfill_official_economy_history import ROOT, UA, load_json, save_json, upsert_series, yoy, now_iso, number
from backfill_cpi_iip_history import safe_coverage

URL='https://www.e-stat.go.jp/stat-search/file-download?fileKind=0&statInfId=000040172363'
SHEETS={
  '生産':('industrial_production','鉱工業生産指数'),
  '出荷':('manufacturing_shipments','製造工業出荷指数'),
  '在庫':('industrial_inventories','鉱工業在庫指数'),
  '在庫率':('industrial_inventory_ratio','鉱工業在庫率指数'),
}

def period(v):
    s=str(v or '').strip().replace('.0','')
    m=re.fullmatch(r'((?:19|20)\d{2})(0[1-9]|1[0-2])',s)
    return f'{m.group(1)}-{m.group(2)}' if m else None

def collect():
    r=requests.get(URL,headers={**UA,'Referer':'https://www.e-stat.go.jp/'},timeout=90);r.raise_for_status()
    if not r.content.startswith(b'PK'):raise RuntimeError('e-Stat IIP response is not XLSX')
    wb=load_workbook(io.BytesIO(r.content),read_only=True,data_only=True)
    parsed={}
    for title,(mid,name) in SHEETS.items():
        if title not in wb.sheetnames:raise RuntimeError(f'missing IIP sheet: {title}')
        ws=wb[title]; rows=list(ws.iter_rows(values_only=True))
        header=None; agg=None
        for row in rows[:12]:
            if sum(1 for x in row if period(x))>=24: header=row
            if len(row)>1 and str(row[1] or '').strip()=='鉱工業': agg=row
        if header is None or agg is None:raise RuntimeError(f'header/aggregate row missing: {title}')
        obs=[]
        for c,pv in enumerate(header):
            p=period(pv)
            if not p or c>=len(agg):continue
            v=number(agg[c])
            if v is not None:obs.append({'period':p,'value':round(v,2),'status':'official','source':'経済産業省/e-Stat 鉱工業指数'})
        by={o['period']:o for o in obs};obs=[by[p] for p in sorted(by)]
        if len(obs)<48:raise RuntimeError(f'IIP {title} too short: {len(obs)}')
        parsed[mid]=(name,obs)
    wb.close()
    data=load_json('data/economy/macro.json')
    for mid,(name,obs) in parsed.items():
        yoy(obs);upsert_series(data,mid,name_ja=name,unit='2020=100',observations=obs,basis='seasonally_adjusted_monthly')
    data['historical_backfill_at']=now_iso();save_json('data/economy/macro.json',data)
    return {'status':'success','series':{mid:len(obs) for mid,(_,obs) in parsed.items()},'current_source':URL,'current_official_start':'2018-01','connected_official_target':'1978-01','connected_status':'official METI connected index exists; GitHub direct-download route still pending'}

def main():
    result=collect();p=ROOT/'data/economy/historical-backfill-status.json';status=json.loads(p.read_text(encoding='utf-8'))
    status.setdefault('stages',{})['iip']=result;status['updated_at']=now_iso();status['status']='success' if all(v.get('status') in {'success','delegated'} for v in status['stages'].values()) else 'partial'
    save_json('data/economy/historical-backfill-status.json',status);save_json('data/economy/historical-coverage.json',safe_coverage(status['stages']))
    print(json.dumps(result,ensure_ascii=False,indent=2))
if __name__=='__main__':main()
