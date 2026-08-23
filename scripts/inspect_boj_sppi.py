#!/usr/bin/env python3
import csv, io, re, zipfile
import requests
from openpyxl import load_workbook

UA={'User-Agent':'Mozilla/5.0 LBI-Transport-History/1.0'}
URLS={
 'current':'https://www.stat-search.boj.or.jp/ssi/docs/info/sppi_m_jp.zip',
 'linked':'https://www.stat-search.boj.or.jp/ssi/docs/info/sppilink.csv',
 'list':'https://www.stat-search.boj.or.jp/ssi/docs/info/PRList.xlsx',
}
TERMS=['道路貨物','外航貨物','内航貨物','港湾運送','国際航空貨物','倉庫','3PL','３ＰＬ','サードパーティ']

def dec(b):
    for enc in ('cp932','utf-8-sig','utf-8'):
        try:return b.decode(enc)
        except UnicodeDecodeError:pass
    return b.decode('cp932',errors='replace')

s=requests.Session();s.headers.update(UA)
for k,u in URLS.items():
    r=s.get(u,timeout=90);r.raise_for_status();print(k,'bytes',len(r.content),'ctype',r.headers.get('content-type'))
    if k=='current':
        z=zipfile.ZipFile(io.BytesIO(r.content));print('members',z.namelist())
        for n in z.namelist():
            raw=z.read(n); txt=dec(raw)
            print('FILE',n,'sample',repr(txt[:1200]))
            for term in TERMS:
                hits=[line for line in txt.splitlines() if term.lower() in line.lower()]
                if hits: print(' HIT',term,hits[:5])
    elif k=='linked':
        txt=dec(r.content);print('linked sample',repr(txt[:2500]))
        for term in TERMS:
            hits=[line for line in txt.splitlines() if term.lower() in line.lower()]
            if hits: print(' LINK HIT',term,hits[:5])
    else:
        wb=load_workbook(io.BytesIO(r.content),read_only=True,data_only=True)
        for ws in wb.worksheets:
            found=[]
            for row in ws.iter_rows(values_only=True):
                vals=[str(x or '') for x in row]
                joined=' | '.join(vals)
                if any(t.lower() in joined.lower() for t in TERMS):found.append(vals[:16])
            if found:
                print('SHEET',ws.title,'matches',len(found))
                for row in found[:50]:print(' MAP',row)
        wb.close()
