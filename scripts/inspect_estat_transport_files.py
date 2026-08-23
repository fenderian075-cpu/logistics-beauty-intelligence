#!/usr/bin/env python3
import io,re,requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
H={'User-Agent':'Mozilla/5.0 LBI-Transport-History/1.0','Referer':'https://www.e-stat.go.jp/'}
PORT_LIST='https://www.e-stat.go.jp/stat-search/files?cycle=0&layout=dataset&page=1&tclass1val=0&toukei=00600280&tstat=000001135203'
r=requests.get(PORT_LIST,headers=H,timeout=90);r.raise_for_status();soup=BeautifulSoup(r.content,'html.parser')
books={}
for a in soup.find_all('a',href=True):
    txt=' '.join(a.stripped_strings)
    m=re.search(r'港別集計値\s*(20\d{2})年1月[～~](\d{1,2})月',txt)
    sid=re.search(r'stat_infid=(\d+)',a['href'])
    if m and sid:books[int(m.group(1))]=(int(m.group(2)),sid.group(1))
print('PORT YEARS',sorted(books))
for y in (min(books),2019,2020,max(books)):
    if y not in books:continue
    endm,sid=books[y];url=f'https://www.e-stat.go.jp/stat-search/file-download?statInfId={sid}&fileKind=0'
    rr=requests.get(url,headers=H,timeout=120);rr.raise_for_status();print('\n=== PORT',y,'months',endm,'sid',sid,'bytes',len(rr.content))
    wb=load_workbook(io.BytesIO(rr.content),read_only=True,data_only=True)
    ws=next((x for x in wb.worksheets if 'コンテナ個数' in x.title),None);print('sheets',wb.sheetnames,'target',ws.title if ws else None)
    if ws:
        for i,row in enumerate(ws.iter_rows(values_only=True)):
            vals=tuple(row[:45])
            if any(v not in (None,'') for v in vals):print(i+1,repr(vals))
            if i>=32:break
    wb.close()
