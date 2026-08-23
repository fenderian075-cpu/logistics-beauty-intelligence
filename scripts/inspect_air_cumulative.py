#!/usr/bin/env python3
import io,requests
from openpyxl import load_workbook
URL='https://www.e-stat.go.jp/stat-search/file-download?fileKind=0&statInfId=000040471594'
H={'User-Agent':'Mozilla/5.0 LBI-Air-History/1.0','Referer':'https://www.e-stat.go.jp/'}
r=requests.get(URL,headers=H,timeout=90);r.raise_for_status();print('bytes',len(r.content),r.headers.get('content-type'))
if not r.content.startswith(b'PK'):raise SystemExit(repr(r.content[:200]))
wb=load_workbook(io.BytesIO(r.content),read_only=True,data_only=True);print('sheets',wb.sheetnames)
for ws in wb.worksheets:
    print('\nSHEET',ws.title,ws.max_row,ws.max_column)
    non=[]
    for i,row in enumerate(ws.iter_rows(values_only=True)):
        vals=tuple(row[:40])
        if any(v not in (None,'') for v in vals):non.append((i+1,vals))
    for x in non[:80]:print(x[0],repr(x[1]))
    print('TAIL')
    for x in non[-30:]:print(x[0],repr(x[1]))
wb.close()
