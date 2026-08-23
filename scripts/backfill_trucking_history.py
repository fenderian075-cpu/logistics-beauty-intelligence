#!/usr/bin/env python3
"""Backfill Japan commercial-truck tonnes and tonne-km from MLIT/e-Stat.

The monthly survey publishes rolling summary workbooks 2-1 and 2-2.  January
snapshots contain the preceding calendar year's twelve monthly observations,
so one snapshot per year reconstructs the monthly history with far fewer
requests than downloading every monthly 3-1 table.

MLIT notes that pre-April-2020 values in these trend tables were retrospectively
adjusted with connection coefficients after the survey-method change. We retain
those official connected values as published; no interpolation is performed.
"""
from __future__ import annotations

import io,json,re
from datetime import datetime,timezone,timedelta
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
OUT=ROOT/'data/economy/trucking.json'
STATUS=ROOT/'data/economy/trucking-history-status.json'
BASE='https://www.e-stat.go.jp'
LIST='https://www.e-stat.go.jp/stat-search/files?cycle=1&layout=dataset&page={page}&tclass1val=0&toukei=00600330&tstat=000001017236'
H={'User-Agent':'Mozilla/5.0 LBI-Trucking-History/1.0','Referer':'https://www.e-stat.go.jp/'}
JST=timezone(timedelta(hours=9))

def now():return datetime.now(JST).replace(microsecond=0).isoformat()

def month_from_context(text):
    m=re.search(r'(20\d{2})年\s*(\d{1,2})月分',text)
    return (int(m.group(1)),int(m.group(2))) if m else None

def discover(session):
    found={'2-1':{},'2-2':{}}
    empty=0
    for page in range(1,70):
        r=session.get(LIST.format(page=page),timeout=90);r.raise_for_status();soup=BeautifulSoup(r.content,'html.parser')
        added=0
        for a in soup.find_all('a',href=True):
            text=' '.join(a.stripped_strings);parent=' '.join(a.parent.stripped_strings) if a.parent else text
            context=text+' '+parent
            table=None
            if '2-1' in text or '２－１' in text or '輸送トン数の推移' in text:table='2-1'
            elif '2-2' in text or '２－２' in text or '輸送トンキロの推移' in text:table='2-2'
            if not table:continue
            sid=re.search(r'stat_infid=(\d+)',a.get('href',''))
            ym=month_from_context(context)
            if not sid or not ym:continue
            key=f'{ym[0]:04d}-{ym[1]:02d}'
            if key not in found[table]:
                found[table][key]=f'{BASE}/stat-search/file-download?statInfId={sid.group(1)}&fileKind=0';added+=1
        empty=empty+1 if added==0 else 0
        years={int(k[:4]) for d in found.values() for k in d}
        if years and min(years)<=2016 and max(years)>=2026 and empty>=2:break
    return found

def parse_time_code(v):
    s=str(v or '').strip().replace('.0','')
    m=re.match(r'^(20\d{2})(\d{2})',s)
    if m:
        y,mo=int(m.group(1)),int(m.group(2))
        if 1<=mo<=12:return f'{y:04d}-{mo:02d}'
    return None

def parse_summary(blob, sheet, source_status):
    wb=load_workbook(io.BytesIO(blob),read_only=True,data_only=True);ws=wb[sheet]
    out=[]
    for row in ws.iter_rows(values_only=True):
        if len(row)<5:continue
        p=parse_time_code(row[1])
        if not p:continue
        try:x=float(row[4])
        except (TypeError,ValueError):continue
        out.append({'period':p,'value':x,'status':source_status,'source':'国土交通省 自動車輸送統計/e-Stat'})
    wb.close();return out

def changes(obs):
    by={r['period']:r['value'] for r in obs}
    for i,r in enumerate(obs):
        if i:
            pv=obs[i-1]['value']
            if pv:r['mom']=round((r['value']/pv-1)*100,2)
        y,m=map(int,r['period'].split('-'));pv=by.get(f'{y-1:04d}-{m:02d}')
        if pv:r['yoy']=round((r['value']/pv-1)*100,2)

def collect_metric(session,files,sheet):
    # Prefer January snapshots: they carry the prior full year's monthly history.
    picks=sorted((k,u) for k,u in files.items() if k.endswith('-01'))
    if files:
        latest=max(files)
        if latest not in {k for k,_ in picks}:picks.append((latest,files[latest]))
    merged={};used=[];fail={}
    for snap,url in picks:
        try:
            r=session.get(url,timeout=90);r.raise_for_status()
            if not r.content.startswith(b'PK'):raise RuntimeError('not xlsx')
            status='official_connected' if snap<'2021-01' else 'official'
            rows=parse_summary(r.content,sheet,status)
            for x in rows:
                # summary workbook includes annual rows too, but parser only accepts YYYYMM.
                merged[x['period']]=x
            used.append(snap)
        except Exception as e:fail[snap]=f'{type(e).__name__}: {e}'
    rows=[merged[p] for p in sorted(merged)];changes(rows);return rows,used,fail

def upsert(data,mid,name,unit,obs):
    s=next((x for x in data.setdefault('series',[]) if x.get('metric_id')==mid),None)
    if s is None:s={'metric_id':mid};data['series'].append(s)
    s.update({'name_ja':name,'unit':unit,'observations':obs})

def main():
    s=requests.Session();s.headers.update(H);files=discover(s)
    ton,used1,fail1=collect_metric(s,files['2-1'],'2-1');tkm,used2,fail2=collect_metric(s,files['2-2'],'2-2')
    if len(ton)<100 or len(tkm)<100:
        raise RuntimeError(f'truck history too short tonnes={len(ton)} tonkm={len(tkm)} discovered={{k:len(v) for k,v in files.items()}} fail={fail1|fail2}')
    if ton[0]['period']>'2015-01' or tkm[0]['period']>'2015-01':raise RuntimeError(f'truck history starts too late {ton[0]["period"]}/{tkm[0]["period"]}')
    # Convert official workbook units to existing LBI units.
    ton_out=[{**r,'value':round(r['value']/1000,3)} for r in ton] # thousand t -> million t
    tkm_out=[{**r,'value':round(r['value']/1000,3)} for r in tkm] # million t-km -> billion t-km
    # Recompute % changes after scale conversion is harmless but keeps consistent rounding.
    for rows in (ton_out,tkm_out):
        for r in rows:r.pop('mom',None);r.pop('yoy',None)
        changes(rows)
    data=json.loads(OUT.read_text(encoding='utf-8'))
    upsert(data,'commercial_truck_tonnage','営業用トラック貨物輸送量','million tonnes',ton_out)
    upsert(data,'commercial_truck_ton_km','営業用トラック貨物輸送トンキロ','billion tonne-km',tkm_out)
    data['historical_backfill_at']=now();data['continuity_note']='2020年4月の調査方法変更に伴い、国土交通省公表の推移表は接続係数により2020年3月以前を遡及改訂した系列を使用。'
    OUT.write_text(json.dumps(data,ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
    st={'schema_version':'1.0','dataset':'trucking-history-status','updated_at':now(),'status':'success','discovered':{k:len(v) for k,v in files.items()},'snapshots_used':{'tonnes':used1,'ton_km':used2},'failures':{'tonnes':fail1,'ton_km':fail2},'coverage':{'commercial_truck_tonnage':{'observations':len(ton_out),'start':ton_out[0]['period'],'end':ton_out[-1]['period']},'commercial_truck_ton_km':{'observations':len(tkm_out),'start':tkm_out[0]['period'],'end':tkm_out[-1]['period']}}}
    STATUS.write_text(json.dumps(st,ensure_ascii=False,indent=2)+'\n',encoding='utf-8');print(json.dumps(st,ensure_ascii=False,indent=2))
if __name__=='__main__':main()
