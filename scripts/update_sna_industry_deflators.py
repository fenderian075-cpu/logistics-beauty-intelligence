#!/usr/bin/env python3
"""Collect Japan SNA economic-activity nominal, real and GDP-deflator history.

Canonical source: Cabinet Office ESRI annual SNA Excel files.
The collector intentionally downloads all three tables from the same annual vintage
and emits a same-period decomposition dataset. CPI/SPPI are never substituted for
SNA GDP deflators.
"""
from __future__ import annotations

import argparse
import io
import json
import math
import re
import statistics
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import requests
from openpyxl import load_workbook

JST = timezone(timedelta(hours=9))
DEFAULT_VINTAGE = 2024
URLS = {
    "nominal": "https://www.esri.cao.go.jp/jp/sna/data/data_list/kakuhou/files/{v}/tables/{v}fcm3n_jp.xlsx",
    "real": "https://www.esri.cao.go.jp/jp/sna/data/data_list/kakuhou/files/{v}/tables/{v}fcm3rn_jp.xlsx",
    "deflator": "https://www.esri.cao.go.jp/jp/sna/data/data_list/kakuhou/files/{v}/tables/{v}fcm3dn_jp.xlsx",
}

INDUSTRIES = {
    "agriculture": ["農林水産業"],
    "mining": ["鉱業"],
    "manufacturing": ["製造業"],
    "utilities_waste": ["電気・ガス・水道・廃棄物処理業", "電気・ガス・水道・廃棄物処理"],
    "construction": ["建設業"],
    "wholesale_retail": ["卸売・小売業", "卸売・小売"],
    "transport_postal": ["運輸・郵便業", "運輸・郵便"],
    "accommodation_food": ["宿泊・飲食サービス業", "宿泊・飲食サービス"],
    "information_communications": ["情報通信業", "情報通信"],
    "finance_insurance": ["金融・保険業", "金融・保険"],
    "real_estate": ["不動産業", "不動産"],
    "professional_business_support": ["専門・科学技術、業務支援サービス業", "専門・科学技術、業務支援サービス"],
    "public_admin": ["公務"],
    "education": ["教育"],
    "health_social": ["保健衛生・社会事業", "保健衛生・社会事業"],
    "other_services": ["その他のサービス"],
}

JP_NAMES = {
    "agriculture": "農林水産業", "mining": "鉱業", "manufacturing": "製造業",
    "utilities_waste": "電気・ガス・水道・廃棄物処理業", "construction": "建設業",
    "wholesale_retail": "卸売・小売業", "transport_postal": "運輸・郵便業",
    "accommodation_food": "宿泊・飲食サービス業", "information_communications": "情報通信業",
    "finance_insurance": "金融・保険業", "real_estate": "不動産業",
    "professional_business_support": "専門・科学技術、業務支援サービス業",
    "public_admin": "公務", "education": "教育", "health_social": "保健衛生・社会事業",
    "other_services": "その他のサービス",
}


def norm(value) -> str:
    if value is None:
        return ""
    s = str(value).replace("\u3000", " ")
    s = re.sub(r"\s+", "", s)
    s = s.replace("（", "(").replace("）", ")")
    s = re.sub(r"^[0-9０-９]+[\.．、]\s*", "", s)
    return s


def year_of(value) -> Optional[int]:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        n = int(value)
        return n if 1900 <= n <= 2100 else None
    m = re.search(r"(?<!\d)(19\d{2}|20\d{2})(?!\d)", str(value))
    return int(m.group(1)) if m else None


def num(value) -> Optional[float]:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)
    s = str(value).strip().replace(",", "")
    if s in {"", "-", "…", "..", "—", "―"}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def industry_id(value) -> Optional[str]:
    s = norm(value)
    if not s:
        return None
    for iid, aliases in INDUSTRIES.items():
        for alias in aliases:
            a = norm(alias)
            if s == a or s.endswith(a):
                return iid
    return None


def download(url: str) -> bytes:
    r = requests.get(url, timeout=60, headers={"User-Agent": "LBI-SNA-Collector/1.0"})
    r.raise_for_status()
    if len(r.content) < 5000:
        raise RuntimeError(f"Downloaded file is unexpectedly small: {url}")
    return r.content


def horizontal_candidates(wb) -> List[Dict[str, Dict[int, float]]]:
    """Extract matrices where years are columns and industries are rows."""
    out = []
    for ws in wb.worksheets:
        max_col = min(ws.max_column, 120)
        max_row = min(ws.max_row, 500)
        header_rows = []
        for r in range(1, max_row + 1):
            years = {}
            for c in range(1, max_col + 1):
                y = year_of(ws.cell(r, c).value)
                if y is not None:
                    years[c] = y
            if len(set(years.values())) >= 5:
                header_rows.append((r, years))
        for hr, year_cols in header_rows:
            data: Dict[str, Dict[int, float]] = {}
            for r in range(hr + 1, min(max_row, hr + 120) + 1):
                iid = None
                for c in range(1, min(max_col, 18) + 1):
                    iid = industry_id(ws.cell(r, c).value)
                    if iid:
                        break
                if not iid:
                    continue
                vals = {}
                for c, y in year_cols.items():
                    v = num(ws.cell(r, c).value)
                    if v is not None:
                        vals[y] = v
                if len(vals) >= 3:
                    data[iid] = vals
            if len(data) >= 8:
                out.append(data)
    return out


def vertical_candidates(wb) -> List[Dict[str, Dict[int, float]]]:
    """Extract transposed matrices where years are rows and industries are columns."""
    out = []
    for ws in wb.worksheets:
        max_col = min(ws.max_column, 120)
        max_row = min(ws.max_row, 500)
        # Find rows with many industry labels; they are likely column headers.
        for hr in range(1, max_row + 1):
            industry_cols = {}
            for c in range(1, max_col + 1):
                iid = industry_id(ws.cell(hr, c).value)
                if iid:
                    industry_cols[c] = iid
            if len(industry_cols) < 8:
                continue
            data: Dict[str, Dict[int, float]] = {iid: {} for iid in industry_cols.values()}
            for r in range(hr + 1, min(max_row, hr + 120) + 1):
                y = None
                for c in range(1, min(max_col, 12) + 1):
                    y = year_of(ws.cell(r, c).value)
                    if y is not None:
                        break
                if y is None:
                    continue
                for c, iid in industry_cols.items():
                    v = num(ws.cell(r, c).value)
                    if v is not None:
                        data[iid][y] = v
            data = {k: v for k, v in data.items() if len(v) >= 3}
            if len(data) >= 8:
                out.append(data)
    return out


def score_candidate(data: Dict[str, Dict[int, float]], kind: str) -> float:
    years = sum(len(v) for v in data.values())
    score = len(data) * 100 + years
    if kind == "deflator":
        vals2020 = [series.get(2020) for series in data.values() if series.get(2020) is not None]
        if vals2020:
            near = sum(1 for v in vals2020 if 95 <= v <= 105)
            score += near * 500
            med = statistics.median(vals2020)
            score -= abs(med - 100) * 20
    else:
        vals = [v for series in data.values() for y, v in series.items() if y >= 2020]
        if vals and statistics.median(vals) > 500:
            score += 3000
    return score


def extract_table(blob: bytes, kind: str) -> Dict[str, Dict[int, float]]:
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    candidates = horizontal_candidates(wb) + vertical_candidates(wb)
    if not candidates:
        raise RuntimeError(f"No SNA matrix found in {kind} workbook")
    best = max(candidates, key=lambda d: score_candidate(d, kind))
    missing = sorted(set(INDUSTRIES) - set(best))
    if len(best) < 14:
        raise RuntimeError(f"Only {len(best)} industries parsed from {kind}; missing={missing}")
    if kind == "deflator":
        vals2020 = [s.get(2020) for s in best.values() if s.get(2020) is not None]
        if len(vals2020) < 10 or sum(1 for v in vals2020 if 95 <= v <= 105) < 8:
            raise RuntimeError("Deflator sanity check failed: 2020 should be approximately 100")
    return best


def growth(series: Dict[int, float], year: int) -> Optional[float]:
    a, b = series.get(year - 1), series.get(year)
    if a in (None, 0) or b is None:
        return None
    return ((b / a) - 1) * 100


def round1(value: Optional[float]) -> Optional[float]:
    return None if value is None else round(value, 1)


def build(vintage: int) -> dict:
    blobs = {k: download(url.format(v=vintage)) for k, url in URLS.items()}
    tables = {k: extract_table(blob, k) for k, blob in blobs.items()}
    common_years = sorted(set.intersection(*[
        set(y for s in table.values() for y in s) for table in tables.values()
    ]))
    if not common_years or common_years[-1] < vintage:
        raise RuntimeError(f"Latest common year is not {vintage}: {common_years[-5:] if common_years else []}")

    industries = []
    for iid in INDUSTRIES:
        n = tables["nominal"].get(iid, {})
        r = tables["real"].get(iid, {})
        d = tables["deflator"].get(iid, {})
        years = sorted(set(n) & set(r) & set(d))
        obs = []
        for y in years:
            nominal_yoy = growth(n, y)
            real_yoy = growth(r, y)
            deflator_yoy = growth(d, y)
            obs.append({
                "period": str(y),
                "nominal_gdp_jpy_billion": round(n[y], 1),
                "real_gdp_chain_jpy_billion": round(r[y], 1),
                "deflator_index_2020_100": round(d[y], 1),
                "nominal_yoy_pct": round1(nominal_yoy),
                "real_yoy_pct": round1(real_yoy),
                "deflator_yoy_pct": round1(deflator_yoy),
                "identity_gap_pctpt": round1(None if nominal_yoy is None or real_yoy is None or deflator_yoy is None else nominal_yoy - real_yoy - deflator_yoy),
                "status": "official_levels_with_derived_growth"
            })
        industries.append({"id": iid, "name_ja": JP_NAMES[iid], "observations": obs})

    return {
        "schema_version": "1.0",
        "dataset": "industry-deflators",
        "title_ja": "SNA産業別 名目・実質・GDPデフレーター",
        "frequency": "annual",
        "source_vintage": str(vintage),
        "updated_at": datetime.now(JST).replace(microsecond=0).isoformat(),
        "source": {
            "name": "内閣府 国民経済計算年次推計 主要系列表 3. 経済活動別国内総生産",
            "landing_url": f"https://www.esri.cao.go.jp/jp/sna/data/data_list/kakuhou/files/{vintage}/{vintage}_kaku_top.html",
            "nominal_url": URLS["nominal"].format(v=vintage),
            "real_url": URLS["real"].format(v=vintage),
            "deflator_url": URLS["deflator"].format(v=vintage),
        },
        "methodology": "Nominal, real and deflator levels are collected from the same Cabinet Office annual SNA vintage. YoY fields are derived from adjacent official levels. Chain-linked real levels are used only to derive within-series growth; CPI/SPPI are never substituted for the SNA industry GDP deflator.",
        "coverage": {"start_year": min(common_years), "end_year": max(common_years), "target_industries": len(INDUSTRIES)},
        "industries": industries,
    }


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--vintage", type=int, default=DEFAULT_VINTAGE)
    p.add_argument("--output", default="data/economy/industry-deflators.json")
    p.add_argument("--check-only", action="store_true")
    args = p.parse_args()
    data = build(args.vintage)
    if args.check_only:
        print(json.dumps(data["coverage"], ensure_ascii=False))
        return
    path = Path(args.output)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    latest = {row["id"]: row["observations"][-1] for row in data["industries"] if row["observations"]}
    print(f"Wrote {path}: {len(latest)} industries, {data['coverage']}")


if __name__ == "__main__":
    main()
