#!/usr/bin/env python3
"""Backfill MLIT 21-company ordinary-warehouse history from official workbooks.

The official monthly workbook contains a long-run `推移表` with inbound volume,
inventory balance/value, utilization and cargo-turnover ratio.  Outbound volume
is reconstructed only where the workbook provides all components of MLIT's
published turnover formula, and is explicitly marked as derived.
"""
from __future__ import annotations

from io import BytesIO
import json
import re
import unicodedata
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data/economy/warehouse-flow.json"
PAGE = "https://www.mlit.go.jp/seisakutokatsu/freight/seisakutokatsu_freight_mn2_000009.html"
UA = {"User-Agent": "Mozilla/5.0 LBI-Warehouse-History/1.0"}


def norm(value) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).replace(" ", "")


def period_from_label(value, current_year=None):
    text = norm(value)
    if not text:
        return None, current_year
    m = re.search(r"平成(\d+)年(\d+)月", text)
    if m:
        y, mo = 1988 + int(m.group(1)), int(m.group(2)); return f"{y:04d}-{mo:02d}", y
    m = re.search(r"令和(\d+)年(\d+)月", text)
    if m:
        y, mo = 2018 + int(m.group(1)), int(m.group(2)); return f"{y:04d}-{mo:02d}", y
    m = re.search(r"H(\d+)年(\d+)月", text, re.I)
    if m:
        y, mo = 1988 + int(m.group(1)), int(m.group(2)); return f"{y:04d}-{mo:02d}", y
    m = re.search(r"R(\d+)年(\d+)月", text, re.I)
    if m:
        y, mo = 2018 + int(m.group(1)), int(m.group(2)); return f"{y:04d}-{mo:02d}", y
    m = re.fullmatch(r"(\d{1,2})月", text)
    if m and current_year:
        return f"{current_year:04d}-{int(m.group(1)):02d}", current_year
    return None, current_year


def num(value):
    try:
        n = float(value)
        return n if n == n else None
    except (TypeError, ValueError):
        return None


def workbook_links():
    res = requests.get(PAGE, headers=UA, timeout=60); res.raise_for_status()
    soup = BeautifulSoup(res.text, "html.parser")
    links = []
    for a in soup.find_all("a", href=True):
        href = urljoin(PAGE, a["href"])
        if re.search(r"\.xlsx(?:$|\?)", href, re.I): links.append(href)
    # Recent files are xlsx. The page lists months within the latest year first;
    # inspecting the first 24 is enough to select the workbook with the latest
    # embedded long-run trend table without downloading the full archive.
    return list(dict.fromkeys(links))[:24]


def parse_xlsx(url):
    res = requests.get(url, headers=UA, timeout=60); res.raise_for_status()
    book = load_workbook(BytesIO(res.content), data_only=True, read_only=True)
    if "推移表" not in book.sheetnames:
        return None
    ws = book["推移表"]
    rows = []
    current_year = None
    # The monthly section starts after row 30. Columns are fixed in the official
    # workbook: B inbound kt, H inventory kt, K inventory value JPY mn,
    # Q utilization %, R turnover %.
    for r in range(31, ws.max_row + 1):
        period, current_year = period_from_label(ws.cell(r, 1).value, current_year)
        if not period:
            continue
        inbound_kt = num(ws.cell(r, 2).value)
        inventory_kt = num(ws.cell(r, 8).value)
        inventory_value_mn = num(ws.cell(r, 11).value)
        utilization = num(ws.cell(r, 17).value)
        turnover = num(ws.cell(r, 18).value)
        if inbound_kt is None or inventory_kt is None:
            continue
        rows.append({
            "period": period,
            "inbound_kt": inbound_kt,
            "inventory_kt": inventory_kt,
            "inventory_value_mn_jpy": inventory_value_mn,
            "utilization_pct": utilization,
            "turnover_pct": turnover,
        })
    if not rows:
        return None
    # Deduplicate the legacy anchor row and sort chronologically.
    by_period = {row["period"]: row for row in rows}
    rows = [by_period[k] for k in sorted(by_period)]
    return {"url": url, "rows": rows, "latest": rows[-1]["period"]}


def obs(period, value, status="official", **extra):
    row = {"period": period, "value": round(float(value), 3), "status": status,
           "source": "国土交通省 営業普通倉庫の実績（主要21社）"}
    row.update(extra)
    return row


def add_changes(rows):
    values = {r["period"]: r for r in rows}
    for i, row in enumerate(rows):
        if i:
            prev = rows[i-1]["value"]
            if prev not in (None, 0): row["mom"] = round((row["value"] / prev - 1) * 100, 2)
        y, m = map(int, row["period"].split("-"))
        prior = values.get(f"{y-1:04d}-{m:02d}")
        if prior and prior["value"] not in (None, 0): row["yoy"] = round((row["value"] / prior["value"] - 1) * 100, 2)
    return rows


def main():
    parsed = []
    for url in workbook_links():
        try:
            item = parse_xlsx(url)
            if item: parsed.append(item)
        except Exception as exc:
            print("warehouse candidate failed", url, type(exc).__name__, str(exc)[:120])
    if not parsed:
        raise RuntimeError("No usable MLIT warehouse trend workbook found")
    best = max(parsed, key=lambda item: item["latest"])
    raw = best["rows"]

    inbound = [obs(r["period"], r["inbound_kt"] * 1000) for r in raw]
    inventory = [obs(r["period"], r["inventory_kt"] * 1000) for r in raw]
    turnover = [obs(r["period"], r["turnover_pct"], unit_note="percent turnover") for r in raw if r["turnover_pct"] is not None]
    utilization = [obs(r["period"], r["utilization_pct"], unit_note="percent utilization") for r in raw if r["utilization_pct"] is not None]
    inventory_value = [obs(r["period"], r["inventory_value_mn_jpy"] * 1_000_000) for r in raw if r["inventory_value_mn_jpy"] is not None]

    # MLIT formula printed in every workbook:
    # turnover = (inbound + outbound) / (previous inventory + current inventory) * 100.
    # Because turnover is published rounded, outbound is an approximate derived
    # series; never label it as a directly observed official value.
    outbound = []
    for prev, cur in zip(raw, raw[1:]):
        if period_from_label(cur["period"])[0] is None:  # ISO labels do not use this parser; harmless guard below
            pass
        py, pm = map(int, prev["period"].split("-")); cy, cm = map(int, cur["period"].split("-"))
        contiguous = (cy * 12 + cm) - (py * 12 + pm) == 1
        if not contiguous or cur["turnover_pct"] is None:
            continue
        estimated_kt = (cur["turnover_pct"] / 100.0) * (prev["inventory_kt"] + cur["inventory_kt"]) - cur["inbound_kt"]
        if estimated_kt < 0:
            continue
        outbound.append(obs(cur["period"], estimated_kt * 1000,
            status="derived_from_official_turnover_formula",
            derivation="(turnover_pct/100)*(previous_inventory+current_inventory)-inbound",
            note_ja="国交省公表の貨物回転率算式から逆算。公表回転率が丸め値のため概算。"))

    for rows in (inbound, outbound, inventory, turnover, utilization, inventory_value): add_changes(rows)

    data = json.loads(OUT.read_text(encoding="utf-8")) if OUT.exists() else {}
    names = {
        "inbound_volume": ("入庫量", "tonnes", inbound),
        "outbound_volume": ("出庫量（回転率からの推計）", "tonnes", outbound),
        "inventory_balance": ("保管残高", "tonnes", inventory),
        "warehouse_turnover": ("貨物回転率", "pct", turnover),
        "warehouse_utilization": ("倉庫利用率", "pct", utilization),
        "inventory_value": ("保管残高金額", "JPY", inventory_value),
    }
    existing = {s.get("metric_id"): s for s in data.get("series", [])}
    final = []
    for metric_id, (name_ja, unit, rows) in names.items():
        s = existing.get(metric_id, {"metric_id": metric_id})
        s.update({"name_ja": name_ja, "unit": unit, "observations": rows})
        final.append(s)
    # Preserve the legacy storage_revenue placeholder explicitly as unavailable;
    # inventory value is not warehouse service revenue and must not be mislabeled.
    if "storage_revenue" in existing:
        s = existing["storage_revenue"]
        s["observations"] = []
        s["availability_note_ja"] = "主要21社統計に倉庫サービス売上高はないため未取得。保管残高金額とは別概念。"
        final.append(s)
    data.update({
        "schema_version": "1.1",
        "dataset": "warehouse-flow",
        "title_ja": "倉庫荷動き",
        "frequency": "monthly",
        "sources": [{"name": "国土交通省 営業普通倉庫の実績（主要21社）", "url": PAGE, "frequency": "monthly"}],
        "series": final,
        "coverage": {"start": raw[0]["period"], "end": raw[-1]["period"], "observations": len(raw), "workbook": best["url"]},
        "method_note_ja": "入庫・保管残高・利用率・貨物回転率・保管残高金額は公式推移表。出庫量のみ公表回転率算式からの概算派生値。",
    })
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"status":"success","workbook":best["url"],"coverage":data["coverage"],"series":{k:len(v[2]) for k,v in names.items()}}, ensure_ascii=False))

if __name__ == "__main__": main()
