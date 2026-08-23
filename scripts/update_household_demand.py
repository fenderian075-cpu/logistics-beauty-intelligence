#!/usr/bin/env python3
"""Collect national resident-register household counts from e-Stat annual files.

The source is the annual YY-01 total table of the Basic Resident Register
population / vital events / households statistics. No interpolation is used.
Existing verified history is retained; only missing years and the latest year
are fetched on refresh.
"""
from __future__ import annotations

import io
import json
import re
import urllib.parse
import urllib.request
from pathlib import Path

import xlrd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "economy" / "household-demand.json"
PARCEL = ROOT / "data" / "economy" / "parcel-demand.json"
YEARS = list(range(2015, 2027))
LATEST_YEAR = max(YEARS)
LIST_URL = "https://www.e-stat.go.jp/stat-search/files"
DOWNLOAD_URL = "https://www.e-stat.go.jp/stat-search/file-download"
UA = "LBI-official-statistics-collector/1.0"


def fetch_bytes(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def discover_stat_infid(year: int) -> str:
    params = {
        "cycle": "7", "layout": "datalist", "month": "0", "page": "1",
        "result_back": "1", "tclass1": "000001039601", "tclass2val": "0",
        "toukei": "00200241", "tstat": "000001039591", "year": f"{year}0",
    }
    html = fetch_bytes(LIST_URL + "?" + urllib.parse.urlencode(params)).decode("utf-8", errors="replace")
    table_no = f"{str(year)[-2:]}-01"
    candidates = []
    for m in re.finditer(r"stat_infid=(\d+)", html, flags=re.I):
        window = html[max(0, m.start() - 3500):min(len(html), m.end() + 3500)]
        if table_no in window and "人口、人口動態及び世帯数" in window:
            candidates.append(m.group(1))
    if candidates:
        return list(dict.fromkeys(candidates))[0]
    raise RuntimeError(f"could not discover stat_infid for {year} {table_no}")


def as_number(value):
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "").replace("，", "")
        if re.fullmatch(r"-?\d+(?:\.\d+)?", text):
            return float(text)
    return None


def norm(value) -> str:
    return re.sub(r"\s+", "", str(value or "")).replace("　", "")


def workbook_rows(payload: bytes):
    if payload.startswith(b"PK\x03\x04"):
        wb = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
        for ws in wb.worksheets:
            yield ws.title, list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), values_only=True))
        return
    if payload.startswith(b"\xd0\xcf\x11\xe0"):
        book = xlrd.open_workbook(file_contents=payload, on_demand=True)
        for name in book.sheet_names():
            ws = book.sheet_by_name(name)
            yield name, [ws.row_values(i) for i in range(min(ws.nrows, 80))]
        return
    raise RuntimeError(f"unsupported household workbook format; magic={payload[:80]!r}")


def extract_households(payload: bytes, year: int) -> float:
    diagnostics = []
    for title, rows in workbook_rows(payload):
        if not rows:
            continue
        total_rows = [i for i, row in enumerate(rows, start=1) if any(norm(v) == "合計" for v in row[:12])]
        header_cols = set()
        for row in rows[:30]:
            for c_idx, value in enumerate(row, start=1):
                text = norm(value)
                if text == "世帯数" or text.endswith("世帯数"):
                    header_cols.add(c_idx)
        for r_idx in total_rows:
            row = rows[r_idx - 1]
            for c_idx in sorted(header_cols):
                for candidate_col in range(max(1, c_idx - 1), min(len(row), c_idx + 2) + 1):
                    value = as_number(row[candidate_col - 1])
                    if value is not None and 30_000_000 <= value <= 100_000_000:
                        return value
            diagnostics.append((title, r_idx, list(row[:12]), sorted(header_cols)))
    raise RuntimeError(f"household total not found for {year}; diagnostics={diagnostics[:3]}")


def series(data, metric_id):
    for item in data.get("series", []):
        if item.get("metric_id") == metric_id:
            return item
    raise KeyError(metric_id)


def fetch_year(year: int, stat_infid: str | None = None):
    stat_infid = stat_infid or discover_stat_infid(year)
    url = DOWNLOAD_URL + "?" + urllib.parse.urlencode({"fileKind": 0, "statInfId": stat_infid})
    count = extract_households(fetch_bytes(url), year)
    return {
        "period": str(year), "value": round(count / 1_000_000, 6),
        "status": "official_file",
        "source": f"e-Stat BRR {str(year)[-2:]}-01 stat_infid={stat_infid}",
    }, stat_infid


def main():
    data = json.loads(OUT.read_text(encoding="utf-8"))
    parcel = json.loads(PARCEL.read_text(encoding="utf-8"))
    existing = {str(r["period"]): r for r in series(data, "resident_register_households").get("observations", [])}
    discovered = dict(data.get("file_contract", {}).get("stat_infids", {}))

    # Preserve verified history. Fetch only missing years plus the current latest year,
    # so scheduled/PR validation does not redownload the entire archive every run.
    for year in YEARS:
        key = str(year)
        if key in existing and year != LATEST_YEAR:
            continue
        row, stat_infid = fetch_year(year, discovered.get(key))
        existing[key] = row
        discovered[key] = stat_infid

    household_rows = [existing[str(y)] for y in YEARS if str(y) in existing]
    missing = [str(y) for y in YEARS if str(y) not in existing]
    if missing:
        raise RuntimeError(f"missing official household years after refresh: {missing}")

    parcel_values = {str(o["period"]): float(o["value"]) for o in series(parcel, "parcel_delivery_volume").get("observations", [])}
    households = {r["period"]: float(r["value"]) for r in household_rows}
    per_household = []
    for year in sorted(set(parcel_values) & set(households)):
        if "2015" <= year <= "2024" and households[year] > 0:
            per_household.append({
                "period": year,
                "value": round(parcel_values[year] / households[year], 1),
                "status": "derived_proxy",
                "source": "MLIT parcel volume / MIC resident-register households",
            })

    series(data, "resident_register_households")["observations"] = household_rows
    series(data, "parcel_per_household")["observations"] = per_household
    data["status"] = "populated_from_estat_files"
    data["file_contract"] = {"table": "YY-01", "stat_infids": discovered, "refresh_mode": "missing_plus_latest"}
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(json.dumps({
        "status": data["status"],
        "refresh_mode": "missing_plus_latest",
        "household_years": [household_rows[0]["period"], household_rows[-1]["period"]],
        "latest_households_million": household_rows[-1]["value"],
        "parcel_per_household_latest": per_household[-1] if per_household else None,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
