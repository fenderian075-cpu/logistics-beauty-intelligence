#!/usr/bin/env python3
"""Backfill nationwide monthly/annual container throughput from MLIT Port Survey via e-Stat.

Discovers official annual port-aggregation workbooks. Legacy workbooks use
full-width spaces in category labels; newer workbooks do not. In incomplete
current-year files, zero nationwide totals caused by insufficient port replies
are not treated as observations. No interpolation is performed.
"""
from __future__ import annotations

import io,json,re
from datetime import datetime,timezone,timedelta
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
OUT=ROOT/'data/economy/port-throughput.json'
STATUS=ROOT/'data/economy/port-history-status.json'
BASE='https://www.e-stat.go.jp'
LIST='https://www.e-stat.go.jp/stat-search/files?cycle=0&layout=dataset&page={page}&tclass1val=0&toukei=00600280&tstat=000001135203'
H={'User-Agent':'Mozilla/5.0 LBI-Port-History/1.0','Referer':'https://www.e-stat.go.jp/'}
JST=timezone(timedelta(hours=9))

def now():return datetime.now(JST).replace(microsecond=0).isoformat()
def norm(x):return re.sub(r'[\s　]+','',str(x or ''))
def val(x):
    if x in (None,''):return None
    try:return float(str(x).replace(',',''))
    except ValueError:return None

def discover(s):
    found={}
    for page in range(1,6):
        r=s.get(LIST.format(page=page),timeout=90);r.raise_for_status()
        soup=BeautifulSoup(r.content,'html.parser');before=len(found)
        for a in soup.find_all('a',href=True):
            text=' '.join(a.stripped_strings)
            m=re.search(r'港別集計値\s*(20\d{2})年1月[～~](\d{1,2})月',text)
            if not m:continue
            sid=re.search(r'stat_infid=(\d+)',a['href'])
            if not sid:continue
            year=int(m.group(1));end_month=int(m.group(2))
            found[year]={'year':year,'end_month':end_month,'statInfId':sid.group(1),'url':f'{BASE}/stat-search/file-download?statInfId={sid.group(1)}&fileKind=0'}
        if len(found)==before:break
    if min(found,default=9999)>2010 or max(found,default=0)<2025:
        raise RuntimeError(f'Port workbook discovery coverage unexpected: {sorted(found)}')
    return found

def header_row(rows,targets,min_hits=8):
    for i,row in enumerate(rows[:20]):
        hits=sum(norm(x) in targets for x in row)
        if hits>=min_hits:return i
    return None

def total_row(rows,start=0):
    candidates=[]
    for i,row in enumerate(rows[start:start+50],start):
        if any(norm(x)=='合計' for x in row[:3]):
            score=sum((val(x) or 0) for x in row[3:])
            candidates.append((score,i))
    return max(candidates)[1] if candidates else None

def parse_container(blob,meta):
    wb=load_workbook(io.BytesIO(blob),read_only=True,data_only=True)
    ws=next((x for x in wb.worksheets if 'コンテナ個数' in x.title),None)
    if ws is None:raise RuntimeError('コンテナ個数 sheet not found')
    rows=[tuple(r) for r in ws.iter_rows(values_only=True)]
    hi=header_row(rows,{'合計','輸出','輸入','移出','移入'})
    if hi is None:raise RuntimeError(f'container category header not found for {meta["year"]}')
    ri=total_row(rows,hi+1)
    if ri is None:raise RuntimeError(f'container nationwide total row not found for {meta["year"]}')
    header=[norm(x) for x in rows[hi]];data=rows[ri]
    total_cols=[i for i,x in enumerate(header) if x=='合計']
    export_cols=[i for i,x in enumerate(header) if x=='輸出'];import_cols=[i for i,x in enumerate(header) if x=='輸入']
    moveout_cols=[i for i,x in enumerate(header) if x=='移出'];movein_cols=[i for i,x in enumerate(header) if x=='移入']
    n=min(meta['end_month'],len(total_cols),len(export_cols),len(import_cols),len(moveout_cols),len(movein_cols))
    if n<min(meta['end_month'],6):raise RuntimeError(f'container month columns short: {meta}, total={len(total_cols)}')
    out=[]
    for m in range(1,n+1):
        t=val(data[total_cols[m-1]]);ex=val(data[export_cols[m-1]]);im=val(data[import_cols[m-1]]);mo=val(data[moveout_cols[m-1]]);mi=val(data[movein_cols[m-1]])
        if None in (t,ex,im,mo,mi) or t<=0:continue
        out.append({'period':f'{meta["year"]:04d}-{m:02d}','total':round(t),'foreign':round(ex+im),'domestic':round(mo+mi),'export':round(ex),'import':round(im),'status':'official'})
    wb.close();return out

def add_changes(obs):
    by={r['period']:r['value'] for r in obs}
    for i,r in enumerate(obs):
        if i:
            p=obs[i-1]
            if p['value']:r['mom']=round((r['value']/p['value']-1)*100,2)
        y,m=map(int,r['period'].split('-'));pv=by.get(f'{y-1:04d}-{m:02d}')
        if pv:r['yoy']=round((r['value']/pv-1)*100,2)

def upsert(data,mid,unit,obs,name):
    s=next((x for x in data.setdefault('series',[]) if x.get('metric_id')==mid),None)
    if s is None:s={'metric_id':mid};data['series'].append(s)
    s.update({'name_ja':name,'unit':unit,'observations':obs})

def main():
    s=requests.Session();s.headers.update(H);books=discover(s);monthly=[];fail={};partial={}
    for year,meta in sorted(books.items()):
        try:
            r=s.get(meta['url'],timeout=120);r.raise_for_status()
            if not r.content.startswith(b'PK'):raise RuntimeError('not xlsx')
            rows=parse_container(r.content,meta);monthly.extend(rows);print(year,len(rows))
            if len(rows)<meta['end_month']:partial[str(year)]={'published_months':meta['end_month'],'usable_national_months':len(rows)}
        except Exception as e:fail[str(year)]=f'{type(e).__name__}: {e}'
    monthly=sorted({r['period']:r for r in monthly}.values(),key=lambda x:x['period'])
    if not monthly or monthly[0]['period']>'2010-01' or monthly[-1]['period']<'2026-04':
        raise RuntimeError(f'Port monthly coverage too short: {monthly[:1]}..{monthly[-1:]}, failures={fail}')
    specs=[('japan_total_container_monthly','total','全国コンテナ取扱量（月次）'),('japan_foreign_trade_container_monthly','foreign','全国外貿コンテナ取扱量（月次）'),('japan_domestic_container_monthly','domestic','全国内貿コンテナ取扱量（月次）'),('japan_export_container_monthly','export','全国輸出コンテナ取扱量（月次）'),('japan_import_container_monthly','import','全国輸入コンテナ取扱量（月次）')]
    data=json.loads(OUT.read_text(encoding='utf-8'))
    for mid,key,name in specs:
        obs=[{'period':r['period'],'value':r[key],'status':r['status'],'source':'国土交通省 港湾調査/e-Stat'} for r in monthly];add_changes(obs);upsert(data,mid,'TEU',obs,name)
    for mid,key,name in [('japan_total_container','total','全国コンテナ取扱量（年次）'),('japan_foreign_trade_container','foreign','全国外貿コンテナ取扱量（年次）'),('japan_domestic_container','domestic','全国内貿コンテナ取扱量（年次）')]:
        annual=[]
        for y in sorted(set(int(r['period'][:4]) for r in monthly)):
            rs=[r for r in monthly if r['period'].startswith(str(y)+'-')]
            if len(rs)!=12:continue
            annual.append({'period':str(y),'value':round(sum(r[key] for r in rs)/1_000_000,4),'status':'official','source':'国土交通省 港湾調査/e-Stat'})
        by={r['period']:r['value'] for r in annual}
        for r in annual:
            pv=by.get(str(int(r['period'])-1))
            if pv:r['yoy']=round((r['value']/pv-1)*100,2)
        upsert(data,mid,'million TEU',annual,name)
    data['historical_backfill_at']=now();data.setdefault('availability',{})['official_monthly_national_through']=monthly[-1]['period']
    OUT.write_text(json.dumps(data,ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
    status={'schema_version':'1.0','dataset':'port-history-status','updated_at':now(),'status':'success' if not fail else 'partial','workbooks_discovered':len(books),'workbooks_parsed':len(books)-len(fail),'failures':fail,'partial_current_year':partial,'monthly_observations':len(monthly),'start':monthly[0]['period'],'end':monthly[-1]['period']}
    STATUS.write_text(json.dumps(status,ensure_ascii=False,indent=2)+'\n',encoding='utf-8');print(json.dumps(status,ensure_ascii=False,indent=2))
if __name__=='__main__':main()
