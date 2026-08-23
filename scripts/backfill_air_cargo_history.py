#!/usr/bin/env python3
"""Backfill domestic/international air-cargo history from MLIT/e-Stat annual appendix.

The official annual workbook contains calendar-year totals for 2015 onward and
monthly domestic/international observations for the two latest complete years.
Existing newer LBI monthly observations are preserved when the appendix does not
yet cover them. No interpolation is performed.
"""
from __future__ import annotations

import io,json,re
from datetime import datetime,timezone,timedelta
from pathlib import Path

import requests
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
OUT=ROOT/'data/economy/air-cargo.json'
STATUS=ROOT/'data/economy/air-cargo-history-status.json'
URL='https://www.e-stat.go.jp/stat-search/file-download?fileKind=0&statInfId=000040471594'
H={'User-Agent':'Mozilla/5.0 LBI-Air-History/1.0','Referer':'https://www.e-stat.go.jp/'}
JST=timezone(timedelta(hours=9))
SOURCE='国土交通省 航空輸送統計/e-Stat'

def now():return datetime.now(JST).replace(microsecond=0).isoformat()

def number(v):
    if v in (None,''):return None
    s=str(v).strip().replace(',','')
    m=re.match(r'^-?[0-9]+(?:\.[0-9]+)?',s)
    if not m:return None
    try:return float(m.group(0))
    except ValueError:return None

def jp_year(v,current_era=None):
    s=str(v or '').strip()
    if not s:return None,current_era
    if '平成' in s:
        m=re.search(r'平成(\d+)年',s)
        if m:return 1988+int(m.group(1)),'平成'
    if '令和元年' in s:return 2019,'令和'
    if '令和' in s:
        m=re.search(r'令和(\d+)年',s)
        if m:return 2018+int(m.group(1)),'令和'
    m=re.fullmatch(r'(\d+)年',s)
    if m and current_era:
        n=int(m.group(1));return ((1988+n) if current_era=='平成' else (2018+n)),current_era
    m=re.fullmatch(r'(20\d{2})年?',s)
    return (int(m.group(1)),current_era) if m else (None,current_era)

def parse_month(v,year):
    s=str(v or '').strip()
    m=re.search(r'(?:令和(\d+)年\s*)?(\d{1,2})月',s)
    if not m:return None,year
    if m.group(1):year=2018+int(m.group(1))
    if year and 1<=int(m.group(2))<=12:return f'{year:04d}-{int(m.group(2)):02d}',year
    return None,year

def obs(period,value,status='official'):
    return {'period':period,'value':value,'status':status,'source':SOURCE}

def parse_workbook(blob):
    wb=load_workbook(io.BytesIO(blob),read_only=True,data_only=True);ws=wb['付表第1表暦年'] if '付表第1表暦年' in wb.sheetnames else wb[wb.sheetnames[0]]
    rows=[tuple(r) for r in ws.iter_rows(values_only=True)]
    dom_a=[];dom_tkm_a=[];int_a=[];int_tkm_a=[];dom_m=[];dom_tkm_m=[];int_m=[];int_tkm_m=[]

    era=None
    for row in rows[6:17]:
        y,era=jp_year(row[0],era)
        if not y:continue
        w=number(row[5]);tk=number(row[6])
        if w is not None:dom_a.append(obs(str(y),round(w/1000,3)))
        if tk is not None:dom_tkm_a.append(obs(str(y),round(tk/1000,3))) # thousand t-km -> million
    era=None
    for row in rows[50:61]:
        y,era=jp_year(row[8],era)
        if not y:continue
        w=number(row[13]);tk=number(row[14])
        if w is not None:int_a.append(obs(str(y),round(w/1000,3)))
        if tk is not None:int_tkm_a.append(obs(str(y),round(tk/1000,3)))

    year=None
    for row in rows[17:41]:
        p,year=parse_month(row[0],year)
        if not p:continue
        w=number(row[5]);tk=number(row[6])
        if w is not None:dom_m.append(obs(p,round(w/1000,3)))
        if tk is not None:dom_tkm_m.append(obs(p,round(tk/1000,3)))
    year=None
    for row in rows[61:]:
        p,year=parse_month(row[8],year)
        if not p:continue
        w=number(row[13]);tk=number(row[14])
        if w is not None:int_m.append(obs(p,round(w/1000,3)))
        if tk is not None:int_tkm_m.append(obs(p,round(tk/1000,3)))
    wb.close()
    return {'dom_a':dom_a,'dom_tkm_a':dom_tkm_a,'int_a':int_a,'int_tkm_a':int_tkm_a,'dom_m':dom_m,'dom_tkm_m':dom_tkm_m,'int_m':int_m,'int_tkm_m':int_tkm_m}

def merge_newer(existing,new):
    by={r['period']:r for r in new}
    for r in existing or []:
        if r.get('period') and r['period']>max(by,default='0000-00'):
            by[r['period']]=r
    return [by[p] for p in sorted(by)]

def changes(rows,monthly=False):
    by={r['period']:r['value'] for r in rows}
    for i,r in enumerate(rows):
        if monthly and i and rows[i-1]['value']:r['mom']=round((r['value']/rows[i-1]['value']-1)*100,2)
        if monthly:
            y,m=map(int,r['period'].split('-'));prev=by.get(f'{y-1:04d}-{m:02d}')
        else:prev=by.get(str(int(r['period'])-1))
        if prev:r['yoy']=round((r['value']/prev-1)*100,2)

def upsert(data,mid,name,unit,rows,preserve_newer=False):
    s=next((x for x in data.setdefault('series',[]) if x.get('metric_id')==mid),None)
    if s is None:s={'metric_id':mid,'observations':[]};data['series'].append(s)
    merged=merge_newer(s.get('observations',[]),rows) if preserve_newer else rows
    for r in merged:r.pop('mom',None);r.pop('yoy',None)
    changes(merged,'-' in merged[0]['period'] if merged else False)
    s.update({'name_ja':name,'unit':unit,'observations':merged})
    return merged

def main():
    r=requests.get(URL,headers=H,timeout=90);r.raise_for_status()
    if not r.content.startswith(b'PK'):raise RuntimeError('official aviation history is not XLSX')
    p=parse_workbook(r.content)
    for key in ('dom_a','dom_tkm_a','int_a','int_tkm_a'):
        if len(p[key])<11 or p[key][0]['period']>'2015':raise RuntimeError(f'{key} annual history short: {len(p[key])} {p[key][:1]}')
    for key in ('dom_m','dom_tkm_m','int_m','int_tkm_m'):
        if len(p[key])<24:raise RuntimeError(f'{key} monthly history short: {len(p[key])}')

    data=json.loads(OUT.read_text(encoding='utf-8'))
    series={}
    series['dom_m']=upsert(data,'domestic_air_cargo','国内航空貨物輸送量（月次）','thousand tonnes',p['dom_m'],True)
    series['dom_tkm_m']=upsert(data,'domestic_air_cargo_ton_km','国内航空貨物輸送トンキロ（月次）','million tonne-km',p['dom_tkm_m'],True)
    series['int_m']=upsert(data,'international_air_cargo','国際航空貨物輸送量（月次）','thousand tonnes',p['int_m'],True)
    series['int_tkm_m']=upsert(data,'international_air_cargo_ton_km','国際航空貨物輸送トンキロ（月次）','million tonne-km',p['int_tkm_m'],True)
    series['dom_a']=upsert(data,'domestic_air_cargo_annual','国内航空貨物輸送量（年次）','thousand tonnes',p['dom_a'])
    series['dom_tkm_a']=upsert(data,'domestic_air_cargo_ton_km_annual','国内航空貨物輸送トンキロ（年次）','million tonne-km',p['dom_tkm_a'])
    series['int_a']=upsert(data,'international_air_cargo_annual','国際航空貨物輸送量（年次）','thousand tonnes',p['int_a'])
    series['int_tkm_a']=upsert(data,'international_air_cargo_ton_km_annual','国際航空貨物輸送トンキロ（年次）','million tonne-km',p['int_tkm_a'])
    data['historical_backfill_at']=now();OUT.write_text(json.dumps(data,ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
    coverage={k:{'observations':len(v),'start':v[0]['period'],'end':v[-1]['period']} for k,v in series.items()}
    st={'schema_version':'1.0','dataset':'air-cargo-history-status','updated_at':now(),'status':'success','source_url':URL,'coverage':coverage}
    STATUS.write_text(json.dumps(st,ensure_ascii=False,indent=2)+'\n',encoding='utf-8');print(json.dumps(st,ensure_ascii=False,indent=2))

if __name__=='__main__':main()
