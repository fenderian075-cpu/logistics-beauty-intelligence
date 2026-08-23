#!/usr/bin/env python3
import io,requests
from openpyxl import load_workbook
H={'User-Agent':'Mozilla/5.0 LBI-Truck-Diagnostic/1.0','Referer':'https://www.e-stat.go.jp/'}
URLS={
 '2-1':'https://www.e-stat.go.jp/stat-search/file-download?statInfId=000040488230&fileKind=0',
 '2-2':'https://www.e-stat.go.jp/stat-search/file-download?statInfId=000040488231&fileKind=0',
}
for name,url in URLS.items():
    r=requests.get(url,headers=H,timeout=90);r.raise_for_status();print('\n===',name,len(r.content),r.headers.get('content-type'))
    wb=load_workbook(io.BytesIO(r.content),read_only=True,data_only=True);print('sheets',wb.sheetnames)
    for ws in wb.worksheets:
        print('SHEET',ws.title,'rows',ws.max_row,'cols',ws.max_column)
        non=[]
        for i,row in enumerate(ws.iter_rows(values_only=True)):
            vals=tuple(row[:20])
            if any(v not in (None,'') for v in vals):non.append((i+1,vals))
        for i,vals in non[:70]:print(i,repr(vals))
        print('TAIL')
        for i,vals in non[-20:]:print(i,repr(vals))
    wb.close()
