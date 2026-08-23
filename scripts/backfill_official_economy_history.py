#!/usr/bin/env python3
"""Backfill long official Japanese economy/statistics histories used by LBI.

Principles:
- Official public sources only.
- Never fabricate missing observations.
- Each source is isolated: one parser failure must not erase other successful histories.
- Canonical files are updated only after basic semantic/sanity validation.
- Existing machine IDs are preserved.

Sources currently handled:
1. Japan Customs: annual export/import totals (1950-) and world monthly CSV (1979-).
2. Cabinet Office ESRI: latest official GDP time-series CSVs discovered from the SNA menu.
3. Statistics Bureau / e-Stat: CPI long-term national monthly CSVs (1970-, some aggregate series earlier).
4. METI IIP: 2020-base connected monthly indices (1978-) plus current monthly file.

SNA industry nominal/real/deflator history is handled by update_sna_industry_deflators.py
and fuel history by update_japan_fuel_prices_v2.py; this script records their coverage too.
"""
from __future__ import annotations

import csv
import io
import json
import math
import re
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
JST = timezone(timedelta(hours=9))
UA = {"User-Agent": "Mozilla/5.0 LBI-Historical-Backfill/1.0"}

TRADE_ANNUAL_URL = "https://www.customs.go.jp/toukei/suii/html/nenbet.htm"
TRADE_MONTHLY_URL = "https://www.customs.go.jp/toukei/suii/html/data/d41ma.csv"
ESRI_MENU_URL = "https://www.esri.cao.go.jp/jp/sna/menu.html"
CPI_PAGE_URL = (
    "https://www.e-stat.go.jp/stat-search/files?cycle=0&layout=datalist&page=1&"
    "tclass1=000001150151&tclass2=000001150152&tclass3=000001150153&"
    "tclass4=000001150156&tclass5val=0&toukei=00200573&tstat=000001150147"
)
IIP_PAGE_URL = "https://www.meti.go.jp/statistics/tyo/iip/b2020_result-2.html"
IIP_CONNECTED_URL = "https://www.meti.go.jp/statistics/tyo/iip/result/xls/b2020_sgs1j.xlsx"
IIP_CURRENT_URL = "https://www.meti.go.jp/statistics/tyo/iip/result/xls/b2020_gsm1j.xlsx"

STATUS_PATH = ROOT / "data/economy/historical-backfill-status.json"
COVERAGE_PATH = ROOT / "data/economy/historical-coverage.json"


def now_iso() -> str:
    return datetime.now(JST).replace(microsecond=0).isoformat()


def get(url: str, timeout: int = 60) -> requests.Response:
    r = requests.get(url, headers=UA, timeout=timeout)
    r.raise_for_status()
    return r


def load_json(rel: str) -> dict:
    return json.loads((ROOT / rel).read_text(encoding="utf-8"))


def save_json(rel: str, data: dict) -> None:
    path = ROOT / rel
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def decode_bytes(blob: bytes) -> str:
    for enc in ("utf-8-sig", "cp932", "shift_jis", "utf-8"):
        try:
            return blob.decode(enc)
        except UnicodeDecodeError:
            pass
    return blob.decode("utf-8", errors="replace")


def number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        v = float(value)
        return v if math.isfinite(v) else None
    s = str(value).strip().replace(",", "").replace("▲", "-").replace("△", "-")
    s = s.replace("−", "-").replace("―", "-")
    if s in {"", "-", "—", "…", "..", "***", "X"}:
        return None
    s = re.sub(r"[^0-9+\-.]", "", s)
    try:
        return float(s) if s not in {"", "+", "-", "."} else None
    except ValueError:
        return None


def period_token(value: Any) -> str | None:
    s = str(value or "").strip()
    # YYYY-MM / YYYY/M / YYYY年M月
    m = re.search(r"(19\d{2}|20\d{2})\s*(?:[-/年]\s*(\d{1,2})\s*月?)?", s)
    if not m:
        return None
    y = int(m.group(1)); mo = m.group(2)
    return f"{y:04d}-{int(mo):02d}" if mo else f"{y:04d}"


def upsert_series(data: dict, metric_id: str, *, name_ja: str, unit: str, observations: list[dict], **meta: Any) -> None:
    series = data.setdefault("series", [])
    target = next((s for s in series if s.get("metric_id") == metric_id), None)
    if target is None:
        target = {"metric_id": metric_id}
        series.append(target)
    target.update({"name_ja": name_ja, "unit": unit, **meta})
    existing = {str(o.get("period")): o for o in target.get("observations", []) if o.get("period")}
    for obs in observations:
        existing[str(obs["period"])] = obs
    target["observations"] = [existing[k] for k in sorted(existing)]


def yoy(rows: list[dict]) -> list[dict]:
    by_period = {r["period"]: r["value"] for r in rows}
    for r in rows:
        p = r["period"]
        prev = None
        if re.fullmatch(r"\d{4}", p):
            prev = str(int(p) - 1)
        elif re.fullmatch(r"\d{4}-\d{2}", p):
            y, m = map(int, p.split("-")); prev = f"{y-1:04d}-{m:02d}"
        a = by_period.get(prev) if prev else None
        if a not in (None, 0):
            r["yoy"] = round((r["value"] / a - 1) * 100, 2)
    return rows


def collect_trade() -> dict:
    # Annual total 1950- from the official HTML table.
    html = get(TRADE_ANNUAL_URL).text
    soup = BeautifulSoup(html, "html.parser")
    annual = []
    for tr in soup.find_all("tr"):
        cells = [c.get_text(" ", strip=True) for c in tr.find_all(["th", "td"])]
        if len(cells) < 3 or not re.fullmatch(r"19\d{2}|20\d{2}", cells[0]):
            continue
        ex, im = number(cells[1]), number(cells[2])
        if ex is None or im is None:
            continue
        # source unit = thousand JPY; canonical trade file uses trillion JPY.
        annual.append((cells[0], ex / 1_000_000_000, im / 1_000_000_000))
    if len(annual) < 60 or annual[0][0] != "1950":
        raise RuntimeError(f"annual trade parse failed: {len(annual)} rows")

    data = load_json("data/economy/japan-trade.json")
    exp_a = [{"period": p, "value": round(ex, 4), "status": "official", "source": "財務省貿易統計"} for p, ex, _ in annual]
    imp_a = [{"period": p, "value": round(im, 4), "status": "official", "source": "財務省貿易統計"} for p, _, im in annual]
    bal_a = [{"period": p, "value": round(ex-im, 4), "status": "derived_from_official_totals", "source": "財務省貿易統計"} for p, ex, im in annual]
    upsert_series(data, "exports_total_annual", name_ja="輸出額（年次）", unit="JPY_trillion", observations=yoy(exp_a), basis="calendar_year")
    upsert_series(data, "imports_total_annual", name_ja="輸入額（年次）", unit="JPY_trillion", observations=yoy(imp_a), basis="calendar_year")
    upsert_series(data, "trade_balance_annual", name_ja="貿易収支（年次）", unit="JPY_trillion", observations=bal_a, basis="calendar_year")

    monthly_added = 0
    try:
        text = decode_bytes(get(TRADE_MONTHLY_URL).content)
        rows = list(csv.reader(io.StringIO(text)))
        parsed = []
        for row in rows:
            if not row:
                continue
            period = None
            # Common format may split year and month; first try any single cell.
            for cell in row[:6]:
                pt = period_token(cell)
                if pt and "-" in pt:
                    period = pt; break
            if not period and len(row) >= 2:
                y = number(row[0]); m = number(row[1])
                if y and 1979 <= y <= 2100 and m and 1 <= m <= 12:
                    period = f"{int(y):04d}-{int(m):02d}"
            nums = [number(x) for x in row]
            nums = [x for x in nums if x is not None]
            if not period or len(nums) < 2:
                continue
            # World monthly file contains export/import totals; use the two largest monetary fields.
            vals = sorted([x for x in nums if abs(x) > 1_000_000], key=abs, reverse=True)
            if len(vals) >= 2:
                ex, im = vals[0], vals[1]
                parsed.append((period, ex / 1_000_000_000, im / 1_000_000_000))
        # Deduplicate; sanity: official monthly world history starts in 1979 and has hundreds of rows.
        mapp = {p: (ex, im) for p, ex, im in parsed}
        parsed = [(p, *mapp[p]) for p in sorted(mapp)]
        if len(parsed) >= 400 and parsed[0][0] <= "1980-01":
            exp_m = [{"period": p, "value": round(ex, 5), "status": "official", "source": "財務省貿易統計"} for p, ex, _ in parsed]
            imp_m = [{"period": p, "value": round(im, 5), "status": "official", "source": "財務省貿易統計"} for p, _, im in parsed]
            bal_m = [{"period": p, "value": round(ex-im, 5), "status": "derived_from_official_totals", "source": "財務省貿易統計"} for p, ex, im in parsed]
            upsert_series(data, "exports_total", name_ja="輸出額", unit="JPY_trillion", observations=yoy(exp_m), basis="calendar_month")
            upsert_series(data, "imports_total", name_ja="輸入額", unit="JPY_trillion", observations=yoy(imp_m), basis="calendar_month")
            upsert_series(data, "trade_balance", name_ja="貿易収支", unit="JPY_trillion", observations=bal_m, basis="calendar_month")
            monthly_added = len(parsed)
    except Exception as exc:
        data.setdefault("collection_notes", []).append(f"world monthly backfill skipped: {type(exc).__name__}: {exc}")

    data["historical_backfill_at"] = now_iso()
    save_json("data/economy/japan-trade.json", data)
    return {"status": "success", "annual_rows": len(annual), "annual_start": annual[0][0], "annual_end": annual[-1][0], "monthly_rows": monthly_added}


GDP_LINK_PATTERNS = {
    "real_gdp_qoq_pct": re.compile(r"ritu-jk.*\.csv$", re.I),
    "nominal_gdp_qoq_pct": re.compile(r"ritu-mk.*\.csv$", re.I),
    "real_gdp_quarterly_level_saar": re.compile(r"gaku-jk.*\.csv$", re.I),
    "nominal_gdp_quarterly_level_saar": re.compile(r"gaku-mk.*\.csv$", re.I),
    "real_gdp_fy_level": re.compile(r"gaku-jfy.*\.csv$", re.I),
    "nominal_gdp_fy_level": re.compile(r"gaku-mfy.*\.csv$", re.I),
}


def parse_esri_csv(blob: bytes, metric_id: str) -> list[dict]:
    text = decode_bytes(blob)
    rows = list(csv.reader(io.StringIO(text)))
    out = []
    for row in rows:
        if not row:
            continue
        # Period may be in one cell or split Japanese fiscal/quarter labels.
        joined = " ".join(str(x) for x in row[:5])
        period = None
        qm = re.search(r"(19\d{2}|20\d{2}).*?([1-4])[-－〜~ ]*(?:3|6|9|12)月", joined)
        if qm and "quarterly" in metric_id or (qm and "qoq" in metric_id):
            period = f"{qm.group(1)}-Q{qm.group(2)}"
        if not period:
            fy = re.search(r"(19\d{2}|20\d{2})\s*(?:年度|FY)", joined, re.I)
            if fy: period = f"{fy.group(1)}FY"
        if not period:
            # Common ESRI CSV often starts with a date-like cell.
            pt = next((period_token(x) for x in row[:5] if period_token(x)), None)
            if pt:
                period = pt
        if not period:
            continue
        nums = [number(x) for x in row[1:]]
        nums = [x for x in nums if x is not None]
        if not nums:
            continue
        # ESRI time-series CSV puts the main series value toward the right; choose last numeric.
        value = nums[-1]
        if "level" in metric_id:
            # CSV level unit is typically billion/100 million JPY depending table. Convert by magnitude.
            if abs(value) > 10_000:
                value = value / 1000.0  # billion JPY -> trillion JPY
        out.append({"period": period, "value": round(value, 4), "status": "official", "source": "内閣府 国民経済計算"})
    dedup = {r["period"]: r for r in out}
    return [dedup[k] for k in sorted(dedup)]


def collect_gdp() -> dict:
    html = get(ESRI_MENU_URL).text
    soup = BeautifulSoup(html, "html.parser")
    found = {}
    for a in soup.find_all("a", href=True):
        url = urljoin(ESRI_MENU_URL, a["href"])
        for mid, pat in GDP_LINK_PATTERNS.items():
            if pat.search(url): found[mid] = url
    if len(found) < 4:
        raise RuntimeError(f"ESRI time-series links not sufficiently discovered: {sorted(found)}")
    data = load_json("data/economy/macro.json")
    counts = {}
    labels = {
        "real_gdp_qoq_pct": ("実質GDP 前期比", "pct"),
        "nominal_gdp_qoq_pct": ("名目GDP 前期比", "pct"),
        "real_gdp_quarterly_level_saar": ("実質GDP 季節調整済年率", "JPY_trillion"),
        "nominal_gdp_quarterly_level_saar": ("名目GDP 季節調整済年率", "JPY_trillion"),
        "real_gdp_fy_level": ("実質GDP（年度）", "JPY_trillion"),
        "nominal_gdp_fy_level": ("名目GDP（年度）", "JPY_trillion"),
    }
    for mid, url in found.items():
        obs = parse_esri_csv(get(url).content, mid)
        # Do not overwrite with a suspiciously tiny parse.
        minimum = 20 if ("quarterly" in mid or "qoq" in mid) else 10
        if len(obs) < minimum:
            continue
        name, unit = labels[mid]
        upsert_series(data, mid, name_ja=name, unit=unit, observations=obs, source_url=url)
        counts[mid] = len(obs)
    data["historical_backfill_at"] = now_iso()
    save_json("data/economy/macro.json", data)
    if not counts:
        raise RuntimeError("ESRI links found but no series passed parser minimums")
    return {"status": "success", "series": counts, "urls": found}


CPI_LABELS = {
    "cpi_all_items": ["総合"],
    "cpi_food": ["食料"],
    "cpi_housing": ["住居"],
    "cpi_fuel_light_water": ["光熱・水道", "光熱・水道費"],
    "cpi_furniture_household": ["家具・家事用品"],
    "cpi_clothes_footwear": ["被服及び履物", "被服・履物"],
    "cpi_medical": ["保健医療"],
    "cpi_transport_communication": ["交通・通信"],
    "cpi_education": ["教育"],
    "cpi_culture_recreation": ["教養娯楽"],
    "cpi_miscellaneous": ["諸雑費"],
    "cpi_cosmetics": ["化粧品"],
}


def extract_e_stat_csv_links() -> tuple[str, str]:
    soup = BeautifulSoup(get(CPI_PAGE_URL).text, "html.parser")
    middle = item = None
    # e-Stat exposes file-download?fileKind=1&statInfId=... anchors next to titles.
    for a in soup.find_all("a", href=True):
        href = a.get("href", "")
        text = a.get_text(" ", strip=True)
        if "file-download" not in href:
            continue
        parent_text = a.parent.parent.get_text(" ", strip=True) if a.parent and a.parent.parent else text
        url = urljoin(CPI_PAGE_URL, href)
        if "中分類指数" in parent_text and "前月比" not in parent_text and "前年" not in parent_text and middle is None:
            middle = url
        if "品目別価格指数" in parent_text and "前月比" not in parent_text and "前年" not in parent_text and item is None:
            item = url
    # Stable current statInfId fallbacks discovered from the official page.
    middle = middle or "https://www.e-stat.go.jp/stat-search/file-download?fileKind=1&statInfId=000032103842"
    item = item or "https://www.e-stat.go.jp/stat-search/file-download?fileKind=1&statInfId=000032103844"
    return middle, item


def parse_cpi_matrix(blob: bytes, targets: dict[str, list[str]]) -> dict[str, list[dict]]:
    rows = list(csv.reader(io.StringIO(decode_bytes(blob))))
    header_idx = None; columns = {}
    for i, row in enumerate(rows[:80]):
        for c, cell in enumerate(row):
            label = re.sub(r"\s+", "", str(cell))
            for mid, aliases in targets.items():
                if any(re.sub(r"\s+", "", a) == label for a in aliases):
                    columns.setdefault(mid, c)
        if len(columns) >= 4:
            header_idx = i; break
    if header_idx is None:
        return {}
    out = {mid: [] for mid in columns}
    for row in rows[header_idx+1:]:
        if not row: continue
        period = next((period_token(x) for x in row[:6] if period_token(x)), None)
        if not period or "-" not in period: continue
        for mid, c in columns.items():
            if c >= len(row): continue
            v = number(row[c])
            if v is not None and 20 <= v <= 500:
                out[mid].append({"period": period, "value": round(v, 3), "status": "official", "source": "総務省統計局/e-Stat"})
    return {mid: list({r["period"]: r for r in obs}.values()) for mid, obs in out.items() if len(obs) >= 24}


def collect_cpi() -> dict:
    middle_url, item_url = extract_e_stat_csv_links()
    middle = parse_cpi_matrix(get(middle_url).content, {k:v for k,v in CPI_LABELS.items() if k != "cpi_cosmetics"})
    item = parse_cpi_matrix(get(item_url).content, {"cpi_cosmetics": CPI_LABELS["cpi_cosmetics"]})
    parsed = {**middle, **item}
    if len(middle) < 8:
        raise RuntimeError(f"CPI middle-group parser produced only {len(middle)} series")
    data = load_json("data/economy/prices.json")
    names = {s.get("metric_id"): s.get("name_ja") for s in data.get("series", [])}
    for mid, obs in parsed.items():
        obs.sort(key=lambda r:r["period"]); yoy(obs)
        upsert_series(data, mid, name_ja=names.get(mid) or mid, unit="2020=100", observations=obs, basis="national_monthly")
    data["historical_backfill_at"] = now_iso()
    save_json("data/economy/prices.json", data)
    return {"status": "success", "series": {k: len(v) for k,v in parsed.items()}, "middle_url": middle_url, "item_url": item_url}


IIP_LABELS = {
    "industrial_production": ["生産", "生産指数"],
    "manufacturing_shipments": ["出荷", "出荷指数"],
    "industrial_inventories": ["在庫", "在庫指数"],
    "industrial_inventory_ratio": ["在庫率", "在庫率指数"],
}


def parse_iip_workbook(blob: bytes) -> dict[str, list[dict]]:
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    best = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        for hi, row in enumerate(rows[:80]):
            cols = {}
            for c, cell in enumerate(row):
                label = re.sub(r"\s+", "", str(cell or ""))
                for mid, aliases in IIP_LABELS.items():
                    if label in [re.sub(r"\s+", "", a) for a in aliases]: cols[mid] = c
            if len(cols) < 2: continue
            candidate = {mid: [] for mid in cols}
            for r in rows[hi+1:]:
                if not r: continue
                p = next((period_token(x) for x in r[:6] if period_token(x)), None)
                if not p or "-" not in p: continue
                for mid, c in cols.items():
                    if c >= len(r): continue
                    v = number(r[c])
                    if v is not None and 10 <= v <= 500:
                        candidate[mid].append({"period": p, "value": round(v, 2), "status": "official_connected_index", "source": "経済産業省 鉱工業指数"})
            score = sum(len(v) for v in candidate.values())
            if score > sum(len(v) for v in best.values()): best = candidate
    return {k:list({r["period"]:r for r in v}.values()) for k,v in best.items() if len(v)>=24}


def collect_iip() -> dict:
    connected = parse_iip_workbook(get(IIP_CONNECTED_URL).content)
    current = parse_iip_workbook(get(IIP_CURRENT_URL).content)
    merged = {}
    for mid in set(connected) | set(current):
        by = {r["period"]:r for r in connected.get(mid, [])}
        by.update({r["period"]:r for r in current.get(mid, [])})
        merged[mid] = [by[k] for k in sorted(by)]
    if not merged or max(map(len, merged.values())) < 300:
        raise RuntimeError(f"IIP parser did not obtain expected long history: { {k:len(v) for k,v in merged.items()} }")
    data = load_json("data/economy/macro.json")
    for mid, obs in merged.items():
        yoy(obs)
        upsert_series(data, mid, name_ja={
            "industrial_production":"鉱工業生産指数", "manufacturing_shipments":"製造工業出荷指数",
            "industrial_inventories":"鉱工業在庫指数", "industrial_inventory_ratio":"鉱工業在庫率指数"
        }[mid], unit="2020=100", observations=obs, basis="seasonally_adjusted_monthly")
    data["historical_backfill_at"] = now_iso()
    save_json("data/economy/macro.json", data)
    return {"status":"success", "series":{k:len(v) for k,v in merged.items()}, "connected_url":IIP_CONNECTED_URL, "current_url":IIP_CURRENT_URL}


def build_coverage(stage_results: dict) -> dict:
    datasets = []
    for path in sorted((ROOT / "data/economy").glob("*.json")):
        try: data = json.loads(path.read_text(encoding="utf-8"))
        except Exception: continue
        for s in data.get("series", []):
            obs = s.get("observations", []) or []
            periods = sorted(str(o.get("period")) for o in obs if o.get("period") is not None)
            datasets.append({
                "file": str(path.relative_to(ROOT)), "dataset": data.get("dataset"), "metric_id": s.get("metric_id"),
                "name_ja": s.get("name_ja"), "unit": s.get("unit"), "observations": len(periods),
                "start": periods[0] if periods else None, "end": periods[-1] if periods else None,
            })
        # SNA industry dataset has nested industries rather than series.
        for ind in data.get("industries", []):
            obs = ind.get("observations", []) or []
            periods = sorted(str(o.get("period")) for o in obs if o.get("period") is not None)
            datasets.append({
                "file": str(path.relative_to(ROOT)), "dataset": data.get("dataset"), "metric_id": f"industry:{ind.get('id')}",
                "name_ja": ind.get("name_ja"), "unit": "SNA nominal/real/deflator", "observations": len(periods),
                "start": periods[0] if periods else None, "end": periods[-1] if periods else None,
            })
    targets = [
        {"source":"財務省貿易統計", "series":"輸出入総額 年次", "official_start":"1950"},
        {"source":"財務省貿易統計", "series":"輸出入総額 月次", "official_start":"1979-01"},
        {"source":"総務省統計局 CPI", "series":"中分類・品目別 月次", "official_start":"1970-01"},
        {"source":"経済産業省 鉱工業指数", "series":"接続指数 月次", "official_start":"1978-01"},
        {"source":"経済産業省 鉱工業指数", "series":"鉱工業総合 長期", "official_start":"1953"},
        {"source":"内閣府 SNA", "series":"産業別 名目・実質・GDPデフレーター", "official_start":"1994"},
        {"source":"資源エネルギー庁", "series":"給油所小売価格 週次", "official_start":"1990-08-27"},
    ]
    return {"schema_version":"1.0", "dataset":"historical-coverage", "updated_at":now_iso(), "stages":stage_results, "official_targets":targets, "series":datasets}


def main() -> None:
    stages = {}
    for name, func in [("trade", collect_trade), ("gdp", collect_gdp), ("cpi", collect_cpi), ("iip", collect_iip)]:
        try:
            stages[name] = func()
            print(name, "OK", json.dumps(stages[name], ensure_ascii=False))
        except Exception as exc:
            stages[name] = {"status":"error", "error":f"{type(exc).__name__}: {exc}"}
            print(name, "ERROR", stages[name]["error"])
    # Existing long collectors are represented in coverage even when run by the workflow separately.
    stages["sna_industry"] = {"status":"delegated", "collector":"scripts/update_sna_industry_deflators.py"}
    stages["fuel"] = {"status":"delegated", "collector":"scripts/update_japan_fuel_prices_v2.py"}
    status = {"schema_version":"1.0", "dataset":"historical-backfill-status", "updated_at":now_iso(), "status":"success" if all(v.get("status") in {"success","delegated"} for v in stages.values()) else "partial", "stages":stages}
    save_json("data/economy/historical-backfill-status.json", status)
    save_json("data/economy/historical-coverage.json", build_coverage(stages))
    print(json.dumps(status, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
