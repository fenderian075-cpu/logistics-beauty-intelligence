#!/usr/bin/env python3
"""Collect national truck operator and vehicle stock history from MLIT workbooks.

The MLIT "数字でみる自動車" workbooks use both legacy XLS and XLSX and lay
years out vertically by Japanese era. Only explicit official observations are
emitted; no interpolation is performed.
"""
from __future__ import annotations
import io, json, re
from pathlib import Path
import requests, xlrd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / 'data/economy/trucking-physical-capacity.json'
TRUCKING = ROOT / 'data/economy/trucking.json'
SOURCES = {
    'operators': 'https://www.mlit.go.jp/jidosha/content/001894861.xls',
    'vehicles': 'https://www.mlit.go.jp/jidosha/content/001894863.xlsx',
}
HEADERS = {'User-Agent':'Mozilla/5.0 LBI-Trucking-Capacity/1.0','Referer':'https://www.mlit.go.jp/'}
ERA_BASE = {'昭和':1925, '平成':1988, '令和':2018}
ERA_ALIASES = {'昭和':'昭和', '昭':'昭和', '平成':'平成', '平':'平成', '令和':'令和', '令':'令和'}


def norm(v):
    return re.sub(r'\s+', '', str(v or '')).replace('　', '')


def number(v):
    if isinstance(v, (int, float)):
        return float(v)
    s = norm(v).replace(',', '').replace('，', '')
    return float(s) if re.fullmatch(r'-?\d+(?:\.\d+)?', s) else None


def western_year(v):
    s = norm(v).replace('年度', '').replace('年', '')
    if re.fullmatch(r'20\d{2}', s): return int(s)
    m = re.fullmatch(r'(?:H|平成)(\d{1,2})', s, re.I)
    if m: return 1988 + int(m.group(1))
    m = re.fullmatch(r'(?:R|令和)(\d{1,2})', s, re.I)
    if m: return 2018 + int(m.group(1))
    return None


def workbook_rows(payload):
    if payload.startswith(b'PK\x03\x04'):
        wb = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
        for ws in wb.worksheets:
            yield ws.title, [list(r) for r in ws.iter_rows(values_only=True)]
        return
    if payload.startswith(b'\xd0\xcf\x11\xe0'):
        wb = xlrd.open_workbook(file_contents=payload, on_demand=True)
        for name in wb.sheet_names():
            ws = wb.sheet_by_name(name)
            yield name, [ws.row_values(i) for i in range(ws.nrows)]
        return
    raise RuntimeError(f'unsupported workbook magic: {payload[:16]!r}')


def sample_rows(payload, limit=45):
    out=[]
    for sheet, rows in workbook_rows(payload):
        for i, row in enumerate(rows):
            vals=[v for v in row[:18] if norm(v)]
            if vals: out.append((sheet, i+1, vals))
            if len(out) >= limit: return out
    return out


def horizontal_candidates(rows):
    for r_idx, row in enumerate(rows):
        mapping={}
        for c_idx, cell in enumerate(row):
            y=western_year(cell)
            if y and 1989 <= y <= 2035: mapping[y]=c_idx
        if len(mapping) >= 5: yield r_idx, mapping


def parse_horizontal(payload, kind):
    best=None
    for sheet, rows in workbook_rows(payload):
        for yr_idx, cols in horizontal_candidates(rows):
            for r_idx in range(max(0,yr_idx-8), min(len(rows),yr_idx+14)):
                row=rows[r_idx]; label='|'.join(norm(x) for x in row[:10] if norm(x)); score=0
                if kind == 'operators':
                    if '総事業者数' in label: score += 100
                    if '事業者数' in label: score += 40
                    if '合計' in label or '総数' in label: score += 20
                else:
                    if '営業用' in label and '車両' in label: score += 100
                    if '車両数' in label: score += 50
                    if '合計' in label or '総数' in label: score += 20
                if not score: continue
                values={}
                for y,c in cols.items():
                    if c < len(row):
                        v=number(row[c])
                        if v not in (None,0): values[y]=v
                if len(values) >= 5:
                    candidate=(score+len(values),sheet,label,values)
                    if best is None or candidate[0] > best[0]: best=candidate
    if best is None: return None
    _,sheet,label,values=best
    values={y:values[y] for y in sorted(values) if 2010 <= y <= 2035}
    return values, {'layout':'horizontal','sheet':sheet,'matched_label':label}


def decode_era_row(row, current_era):
    era=current_era
    for cell in row[:5]:
        s=norm(cell)
        if not s: continue
        m=re.fullmatch(r'(昭和|昭|平成|平|令和|令)(元|\d{1,2})', s)
        if m:
            era=ERA_ALIASES[m.group(1)]
            return era, 1 if m.group(2) == '元' else int(m.group(2))
        if s in ERA_ALIASES:
            era=ERA_ALIASES[s]
            continue
        if s == '元' and era: return era, 1
        n=number(cell)
        if n is not None and era and 1 <= n <= 64: return era, int(n)
    return era, None


def longest_contiguous(matches):
    """Keep the longest row-ordered block where western years rise by exactly 1.

    This excludes numeric footnotes appearing below the official table body.
    """
    best=[]; current=[]
    for item in matches:
        y=item[0]
        if not current or y == current[-1][0] + 1:
            current.append(item)
        else:
            if len(current) > len(best): best=current
            current=[item]
    if len(current) > len(best): best=current
    return best


def parse_vertical_era(payload, kind):
    best=None
    for sheet, rows in workbook_rows(payload):
        sheet_text='|'.join(norm(x) for row in rows[:8] for x in row[:12] if norm(x))
        if kind == 'operators' and '事業者数' not in sheet_text: continue
        if kind == 'vehicles' and not any(k in sheet_text for k in ('車両数','車両','トラック')): continue

        header_idx=None; header_label=''
        for i,row in enumerate(rows[:25]):
            text='|'.join(norm(x) for x in row[:16] if norm(x))
            if kind == 'operators' and ('合計' in text and ('一般' in text or '特積' in text)):
                header_idx=i; header_label=text; break
            if kind == 'vehicles' and any(k in text for k in ('合計','総数','計')):
                header_idx=i; header_label=text; break
        if header_idx is None:
            for i,row in enumerate(rows[:25]):
                text='|'.join(norm(x) for x in row[:16] if norm(x))
                if '年度' in text: header_idx=i; header_label=text; break
        if header_idx is None: continue

        era=None; matches=[]
        for r_idx,row in enumerate(rows[header_idx+1:], start=header_idx+2):
            era, ey=decode_era_row(row,era)
            if era is None or ey is None: continue
            y=ERA_BASE[era]+ey
            if not (1975 <= y <= 2035): continue
            nums=[number(x) for x in row if number(x) is not None]
            if len(nums) < 2: continue
            total=nums[-1]
            if total <= 0: continue
            matches.append((y,total,r_idx,era,ey))

        block=longest_contiguous(matches)
        modern=[x for x in block if 2010 <= x[0] <= 2035]
        if len(modern) >= 10:
            values={y:total for y,total,_,_,_ in modern}
            last_rows=[(r,era,ey,total) for y,total,r,era,ey in modern[-5:]]
            candidate=(len(modern),sheet,header_label,values,last_rows,(block[0][0],block[-1][0]))
            if best is None or candidate[0] > best[0]: best=candidate

    if best is None: return None
    _,sheet,label,values,last_rows,block_range=best
    return values, {'layout':'vertical_japanese_era','sheet':sheet,'matched_header':label,'official_block_years':list(block_range),'last_rows':last_rows}


def parse_metric(payload, kind):
    parsed=parse_horizontal(payload,kind)
    if parsed and len(parsed[0]) >= 10: return parsed
    parsed=parse_vertical_era(payload,kind)
    if parsed and len(parsed[0]) >= 10: return parsed
    raise RuntimeError(f'{kind}: metric table not found; sample={sample_rows(payload)}')


def obs(values, source):
    rows=[]; prev=None
    for y,value in values.items():
        row={'period':str(y),'value':round(value,3),'status':'official_file','source':source}
        if prev and prev[1]: row['yoy']=round((value/prev[1]-1)*100,2)
        rows.append(row); prev=(y,value)
    return rows


def series(data,mid):
    return next(s for s in data.get('series',[]) if s.get('metric_id') == mid)


def main():
    sess=requests.Session(); sess.headers.update(HEADERS); payloads={}
    for k,u in SOURCES.items():
        r=sess.get(u,timeout=90); r.raise_for_status(); payloads[k]=r.content
    operators,op_meta=parse_metric(payloads['operators'],'operators')
    vehicles,vh_meta=parse_metric(payloads['vehicles'],'vehicles')
    common=sorted(set(operators)&set(vehicles))
    vpo={y:vehicles[y]/operators[y] for y in common if operators[y] > 0}

    trucking=json.loads(TRUCKING.read_text(encoding='utf-8'))
    tkm={int(o['period']):float(o['value']) for o in series(trucking,'commercial_truck_ton_km_annual').get('observations',[])}
    tpv={y:tkm[y]*1_000_000/vehicles[y] for y in sorted(set(tkm)&set(vehicles)) if vehicles[y] > 0}
    if len(vpo) < 8 or len(tpv) < 5:
        raise RuntimeError(f'insufficient overlap: vehicles/operator={sorted(vpo)}, ton-km/vehicle={sorted(tpv)}')

    data=json.loads(OUT.read_text(encoding='utf-8'))
    series(data,'truck_operators')['observations']=obs(operators,'MLIT 数字でみる自動車 トラック事業者数の推移')
    series(data,'commercial_truck_vehicles')['observations']=obs(vehicles,'MLIT 数字でみる自動車 トラック車両数の推移')
    series(data,'vehicles_per_operator')['observations']=[{'period':str(y),'value':round(v,2),'status':'derived','source':'commercial_truck_vehicles / truck_operators'} for y,v in vpo.items()]
    series(data,'ton_km_per_vehicle')['observations']=[{'period':str(y),'value':round(v,1),'status':'derived_proxy','source':'commercial_truck_ton_km_annual / commercial_truck_vehicles'} for y,v in tpv.items()]
    data['status']='populated_from_mlit_files'; data['parser_contract']={'operators':op_meta,'vehicles':vh_meta,'urls':SOURCES}
    OUT.write_text(json.dumps(data,ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
    print(json.dumps({'status':data['status'],'operators':[min(operators),max(operators),operators[max(operators)]],'vehicles':[min(vehicles),max(vehicles),vehicles[max(vehicles)]],'vehicles_per_operator_latest':next(reversed(vpo.values())),'ton_km_per_vehicle_latest':next(reversed(tpv.values())),'parser':data['parser_contract']},ensure_ascii=False,indent=2))

if __name__=='__main__': main()
