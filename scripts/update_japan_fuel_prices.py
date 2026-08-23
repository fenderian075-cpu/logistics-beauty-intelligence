#!/usr/bin/env python3
"""Collect Japan national weekly gasoline/diesel/kerosene retail prices.

Primary source: Agency for Natural Resources and Energy, Petroleum Product Price Survey.
Supports .xlsx/.xls and both common layouts:
1) dates in rows + national price in a column, and
2) product sheets with a `全国` row + dates in columns.
No fixed cell coordinates are used.
"""
from __future__ import annotations

import io, json, re, traceback
from datetime import date, datetime, timezone
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import xlrd

ROOT = Path(__file__).resolve().parents[1]
RESULTS = "https://www.enecho.meti.go.jp/statistics/petroleum_and_lpgas/pl007/results.html"
UA = {"User-Agent": "Mozilla/5.0 LBI-public-fuel-collector/1.3"}
STATUS_PATH = ROOT / "data/economy/fuel-collector-status.json"


def norm(v):
    return re.sub(r"\s+", "", str(v or "")).replace("　", "")


def write_status(status, stage, **extra):
    payload = {"schema_version":"1.2","dataset":"fuel-collector-status","status":status,"stage":stage,
               "updated_at":datetime.now(timezone.utc).isoformat(), **extra}
    STATUS_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def find_weekly_url(html):
    """Choose the long-run 1990- weekly workbook, not the current 29KB detail workbook."""
    soup = BeautifulSoup(html, "html.parser")
    scored = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if not re.search(r"\.xlsx?(?:\?|$)", href, re.I):
            continue
        # The site sometimes places `週次ファイル` outside the anchor itself.
        parent_text = a.parent.get_text(" ", strip=True) if a.parent else ""
        grand_text = a.parent.parent.get_text(" ", strip=True) if a.parent and a.parent.parent else ""
        context = norm("|".join([a.get_text(" ", strip=True), parent_text, grand_text]))
        score = 0
        if "週次ファイル" in context or "週次" in context: score += 100
        if "1990年" in context or "平成2年" in context or "8月27日" in context: score += 80
        if "調査結果一覧" in context: score += 30
        if "結果詳細版" in context or "結果詳細" in context: score -= 120
        if "概要" in context: score -= 60
        scored.append((score, urljoin(RESULTS, href), context[:400]))
    if not scored:
        raise RuntimeError("Excel links not found on official result page")
    scored.sort(key=lambda x: x[0], reverse=True)
    best = scored[0]
    if best[0] <= 0:
        raise RuntimeError(f"long-run weekly workbook could not be identified; candidates={scored[:8]}")
    return best[1], scored[:8]


def parse_date(v, datemode=None):
    if isinstance(v, datetime): return v.date().isoformat()
    if isinstance(v, date): return v.isoformat()
    if datemode is not None and isinstance(v, (int, float)) and 25000 < float(v) < 60000:
        try: return xlrd.xldate_as_datetime(v, datemode).date().isoformat()
        except Exception: pass
    s = str(v or "").strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%Y年%m月%d日"):
        try: return datetime.strptime(s, fmt).date().isoformat()
        except ValueError: pass
    return None


def product_id(text):
    t = norm(text)
    if "レギュラー" in t: return "regular_gasoline_national"
    if "軽油" in t: return "diesel_national"
    if "灯油" in t and "配達" not in t: return "kerosene_national"
    return None


def sheet_blob(name, nrows, ncols, getter):
    cells = [name]
    for r in range(min(nrows, 15)):
        for c in range(min(ncols, 15)):
            v = getter(r, c)
            if v not in (None, ""): cells.append(str(v))
    return "|".join(cells)


def observations_vertical(nrows, ncols, getter, datemode, mid):
    candidate_cols = []
    for c in range(ncols):
        header = "|".join(norm(getter(r, c)) for r in range(min(nrows, 30)))
        if "全国" in header: candidate_cols.append(c)
    best = []
    for col in candidate_cols:
        rows = []
        for r in range(nrows):
            period = None
            for c in range(min(ncols, 8)):
                period = parse_date(getter(r, c), datemode)
                if period: break
            v = getter(r, col)
            if period and isinstance(v, (int, float)) and float(v) > 0:
                rows.append({"period":period,"value":round(float(v),1),"status":"official"})
        if len(rows) > len(best): best = rows
    return best


def observations_horizontal(nrows, ncols, getter, datemode, mid):
    date_cols = {}
    for c in range(ncols):
        for r in range(min(nrows, 30)):
            period = parse_date(getter(r, c), datemode)
            if period:
                date_cols[c] = period
                break
    if len(date_cols) < 2: return []
    best = []
    for r in range(nrows):
        row_label = "|".join(norm(getter(r, c)) for c in range(min(ncols, 12)))
        if "全国" not in row_label: continue
        rows = []
        for c, period in date_cols.items():
            v = getter(r, c)
            if isinstance(v, (int, float)) and float(v) > 0:
                rows.append({"period":period,"value":round(float(v),1),"status":"official"})
        if len(rows) > len(best): best = rows
    return best


def parse_sheet(name, nrows, ncols, getter, datemode=None):
    blob = sheet_blob(name, nrows, ncols, getter)
    mid = product_id(blob)
    if not mid: return None
    vertical = observations_vertical(nrows, ncols, getter, datemode, mid)
    horizontal = observations_horizontal(nrows, ncols, getter, datemode, mid)
    rows = horizontal if len(horizontal) > len(vertical) else vertical
    if not rows: return None
    dedup = {r["period"]: r for r in rows}
    rows = [dedup[p] for p in sorted(dedup)]
    return mid, rows, "horizontal" if len(horizontal) > len(vertical) else "vertical"


def parse_xlsx(content):
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parsed, details = {}, []
    for ws in wb.worksheets:
        result = parse_sheet(ws.title, ws.max_row, ws.max_column, lambda r,c: ws.cell(r+1,c+1).value)
        if result:
            mid, rows, layout = result
            if len(rows) > len(parsed.get(mid, [])): parsed[mid] = rows
            details.append({"sheet":ws.title,"metric_id":mid,"layout":layout,"rows":len(rows)})
    return parsed, details


def parse_xls(content):
    wb = xlrd.open_workbook(file_contents=content)
    parsed, details = {}, []
    for ws in wb.sheets():
        result = parse_sheet(ws.name, ws.nrows, ws.ncols, lambda r,c: ws.cell_value(r,c), wb.datemode)
        if result:
            mid, rows, layout = result
            if len(rows) > len(parsed.get(mid, [])): parsed[mid] = rows
            details.append({"sheet":ws.name,"metric_id":mid,"layout":layout,"rows":len(rows)})
    return parsed, details


def main():
    weekly, link_candidates = None, None
    try:
        write_status("running", "fetch_index")
        page = requests.get(RESULTS, headers=UA, timeout=30); page.raise_for_status()
        weekly, link_candidates = find_weekly_url(page.text)
        write_status("running", "download_workbook", weekly_workbook_url=weekly, link_candidates=link_candidates)
        book = requests.get(weekly, headers=UA, timeout=60); book.raise_for_status(); content = book.content
        if content.startswith(b"PK"):
            workbook_format, (parsed, details) = "xlsx", parse_xlsx(content)
        elif content.startswith(bytes.fromhex("D0CF11E0A1B11AE1")):
            workbook_format, (parsed, details) = "xls", parse_xls(content)
        else:
            raise RuntimeError(f"unsupported workbook signature: {content[:16].hex()}")
        required = ("regular_gasoline_national", "diesel_national")
        if not all(parsed.get(mid) for mid in required):
            raise RuntimeError(f"required national series missing; found={list(parsed)} details={details}")

        path = ROOT / "data/economy/fuel-prices.json"
        data = json.loads(path.read_text(encoding="utf-8")); by_id = {s["metric_id"]:s for s in data.get("series",[])}
        for mid, rows in parsed.items():
            if mid in by_id and rows: by_id[mid]["observations"] = rows
        data["collection_status"] = "official_weekly_history_ingested"
        data["source"]["weekly_workbook_url"] = weekly
        data["last_collected_at"] = datetime.now(timezone.utc).isoformat()
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        counts = {mid:len(rows) for mid,rows in parsed.items()}
        write_status("success","success",weekly_workbook_url=weekly,workbook_format=workbook_format,
                     link_candidates=link_candidates,parsed_sheets=details,row_counts=counts)
        print(counts)
    except Exception as exc:
        write_status("failure","collector",weekly_workbook_url=weekly,link_candidates=link_candidates,
                     error=str(exc),diagnostic_tail=traceback.format_exc()[-6000:])
        raise

if __name__ == "__main__": main()
