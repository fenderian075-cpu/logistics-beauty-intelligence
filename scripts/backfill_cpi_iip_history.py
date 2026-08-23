#!/usr/bin/env python3
"""Harden CPI and IIP long-history backfill against official publication layouts.

CPI e-Stat downloads can be either dates-in-rows or dates-in-columns.
METI IIP file paths are discovered from the official download page rather than guessed.
This script supplements backfill_official_economy_history.py and updates the same
status/coverage diagnostics.
"""
from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from backfill_official_economy_history import (
    ROOT, UA, CPI_PAGE_URL, CPI_LABELS, IIP_PAGE_URL, IIP_LABELS,
    extract_e_stat_csv_links, decode_bytes, number, period_token, upsert_series,
    yoy, load_json, save_json, build_coverage, now_iso,
)


def norm(v) -> str:
    return re.sub(r"\s+", "", str(v or "")).replace("　", "")


def matched_metric(cells, targets):
    norms = [norm(x) for x in cells]
    for mid, aliases in targets.items():
        aset = {norm(a) for a in aliases}
        if any(x in aset for x in norms):
            return mid
    return None


def parse_cpi_any(blob: bytes, targets: dict[str, list[str]]) -> dict[str, list[dict]]:
    rows = list(csv.reader(io.StringIO(decode_bytes(blob))))
    best: dict[str, list[dict]] = {}

    # Layout A: time is in rows, metrics are columns.
    for hi, header in enumerate(rows[:120]):
        cols = {}
        for c, cell in enumerate(header):
            mid = matched_metric([cell], targets)
            if mid:
                cols[mid] = c
        if not cols:
            continue
        candidate = {mid: [] for mid in cols}
        for row in rows[hi + 1:]:
            p = next((period_token(x) for x in row[:10] if period_token(x)), None)
            if not p or "-" not in p:
                continue
            for mid, c in cols.items():
                if c >= len(row):
                    continue
                v = number(row[c])
                if v is not None and 20 <= v <= 500:
                    candidate[mid].append({"period": p, "value": round(v, 3), "status": "official", "source": "総務省統計局/e-Stat"})
        if sum(map(len, candidate.values())) > sum(map(len, best.values())):
            best = candidate

    # Layout B: metrics are rows, time is in columns (common long-history CPI file).
    for hi, header in enumerate(rows[:120]):
        date_cols = {}
        for c, cell in enumerate(header):
            p = period_token(cell)
            if p and "-" in p:
                date_cols[c] = p
        if len(date_cols) < 24:
            continue
        candidate: dict[str, list[dict]] = {}
        for row in rows[hi + 1:]:
            mid = matched_metric(row[:16], targets)
            if not mid:
                continue
            obs = []
            for c, p in date_cols.items():
                if c >= len(row):
                    continue
                v = number(row[c])
                if v is not None and 20 <= v <= 500:
                    obs.append({"period": p, "value": round(v, 3), "status": "official", "source": "総務省統計局/e-Stat"})
            if len(obs) > len(candidate.get(mid, [])):
                candidate[mid] = obs
        if sum(map(len, candidate.values())) > sum(map(len, best.values())):
            best = candidate

    cleaned = {}
    for mid, obs in best.items():
        by = {r["period"]: r for r in obs}
        rows2 = [by[p] for p in sorted(by)]
        if len(rows2) >= 24:
            cleaned[mid] = rows2
    return cleaned


def collect_cpi_hardened() -> dict:
    middle_url, item_url = extract_e_stat_csv_links()
    sess = requests.Session(); sess.headers.update(UA)
    middle_r = sess.get(middle_url, timeout=90); middle_r.raise_for_status()
    item_r = sess.get(item_url, timeout=90); item_r.raise_for_status()
    middle = parse_cpi_any(middle_r.content, {k:v for k,v in CPI_LABELS.items() if k != "cpi_cosmetics"})
    item = parse_cpi_any(item_r.content, {"cpi_cosmetics": CPI_LABELS["cpi_cosmetics"]})
    if len(middle) < 8:
        sample = decode_bytes(middle_r.content)[:3000].replace("\n", " | ")
        raise RuntimeError(f"CPI parser still produced {len(middle)} middle series; sample={sample[:2500]}")
    parsed = {**middle, **item}
    data = load_json("data/economy/prices.json")
    names = {s.get("metric_id"):s.get("name_ja") for s in data.get("series",[])}
    for mid, obs in parsed.items():
        yoy(obs)
        upsert_series(data, mid, name_ja=names.get(mid) or mid, unit="2020=100", observations=obs, basis="national_monthly")
    data["historical_backfill_at"] = now_iso()
    save_json("data/economy/prices.json", data)
    return {"status":"success", "series":{k:len(v) for k,v in parsed.items()}, "middle_url":middle_url, "item_url":item_url}


def discover_iip_links() -> dict[str, str]:
    headers = {**UA, "Accept-Language":"ja,en;q=0.8"}
    r = requests.get(IIP_PAGE_URL, headers=headers, timeout=60); r.raise_for_status()
    soup = BeautifulSoup(r.content, "html.parser")
    wanted = {
        "connected": "b2020_sgs1j.xlsx",
        "current": "b2020_gsm1j.xlsx",
        "aggregate_long": "b2020_sosq1j.xlsx",
    }
    found = {}
    for a in soup.find_all("a", href=True):
        href = a["href"]
        lower = href.lower()
        for key, filename in wanted.items():
            if filename.lower() in lower:
                found[key] = urljoin(IIP_PAGE_URL, href)
    if not all(k in found for k in ("connected","current")):
        raise RuntimeError(f"METI IIP links not found on official page: {found}")
    return found


def iip_metric_from_cells(cells):
    texts = [norm(x) for x in cells]
    for mid, aliases in IIP_LABELS.items():
        aset = {norm(a) for a in aliases}
        if any(t in aset for t in texts):
            return mid
    # Common consolidated table labels.
    joined = "|".join(texts)
    if "生産" in joined and "予測" not in joined and "能力" not in joined: return "industrial_production"
    if "出荷" in joined: return "manufacturing_shipments"
    if "在庫率" in joined: return "industrial_inventory_ratio"
    if "在庫" in joined: return "industrial_inventories"
    return None


def parse_iip_any(blob: bytes) -> dict[str, list[dict]]:
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    best: dict[str, list[dict]] = {}
    for ws in wb.worksheets:
        matrix = [tuple(row) for row in ws.iter_rows(values_only=True)]
        if not matrix: continue

        # Layout A: date rows, metric columns.
        for hi, header in enumerate(matrix[:100]):
            cols = {}
            for c, cell in enumerate(header):
                mid = iip_metric_from_cells([cell])
                if mid: cols[mid] = c
            if not cols: continue
            candidate = {mid:[] for mid in cols}
            for row in matrix[hi+1:]:
                p = next((period_token(x) for x in row[:10] if period_token(x)), None)
                if not p or "-" not in p: continue
                for mid,c in cols.items():
                    if c >= len(row): continue
                    v=number(row[c])
                    if v is not None and 10 <= v <= 500:
                        candidate[mid].append({"period":p,"value":round(v,2),"status":"official_connected_index","source":"経済産業省 鉱工業指数"})
            if sum(map(len,candidate.values())) > sum(map(len,best.values())): best=candidate

        # Layout B: metric rows, date columns.
        for hi, header in enumerate(matrix[:100]):
            date_cols={}
            for c,cell in enumerate(header):
                p=period_token(cell)
                if p and "-" in p: date_cols[c]=p
            if len(date_cols)<12: continue
            candidate={}
            for row in matrix[hi+1:]:
                mid=iip_metric_from_cells(row[:16])
                if not mid: continue
                obs=[]
                for c,p in date_cols.items():
                    if c>=len(row): continue
                    v=number(row[c])
                    if v is not None and 10<=v<=500:
                        obs.append({"period":p,"value":round(v,2),"status":"official_connected_index","source":"経済産業省 鉱工業指数"})
                if len(obs)>len(candidate.get(mid,[])): candidate[mid]=obs
            if sum(map(len,candidate.values())) > sum(map(len,best.values())): best=candidate
    wb.close()
    out={}
    for mid,obs in best.items():
        by={r["period"]:r for r in obs}; rows=[by[p] for p in sorted(by)]
        if len(rows)>=12: out[mid]=rows
    return out


def fetch_meti_xlsx(url, referer):
    headers={**UA,"Referer":referer,"Accept":"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*;q=0.8","Accept-Language":"ja,en;q=0.8"}
    r=requests.get(url,headers=headers,timeout=90); r.raise_for_status()
    if not r.content.startswith(b"PK"):
        raise RuntimeError(f"METI returned non-xlsx content for {url}: {r.content[:80]!r}")
    return r.content


def collect_iip_hardened() -> dict:
    links=discover_iip_links()
    connected=parse_iip_any(fetch_meti_xlsx(links["connected"],IIP_PAGE_URL))
    current=parse_iip_any(fetch_meti_xlsx(links["current"],IIP_PAGE_URL))
    merged={}
    for mid in set(connected)|set(current):
        by={r["period"]:r for r in connected.get(mid,[])}; by.update({r["period"]:r for r in current.get(mid,[])})
        merged[mid]=[by[p] for p in sorted(by)]
    if not merged or max(map(len,merged.values()))<300:
        raise RuntimeError(f"IIP long history too short: { {k:len(v) for k,v in merged.items()} }, links={links}")
    data=load_json("data/economy/macro.json")
    names={"industrial_production":"鉱工業生産指数","manufacturing_shipments":"製造工業出荷指数","industrial_inventories":"鉱工業在庫指数","industrial_inventory_ratio":"鉱工業在庫率指数"}
    for mid,obs in merged.items():
        yoy(obs); upsert_series(data,mid,name_ja=names[mid],unit="2020=100",observations=obs,basis="seasonally_adjusted_monthly")
    data["historical_backfill_at"]=now_iso(); save_json("data/economy/macro.json",data)
    return {"status":"success","series":{k:len(v) for k,v in merged.items()},"links":links}


def main():
    status_path=ROOT/"data/economy/historical-backfill-status.json"
    status=json.loads(status_path.read_text(encoding="utf-8")) if status_path.exists() else {"schema_version":"1.0","dataset":"historical-backfill-status","stages":{}}
    for name,func in (("cpi",collect_cpi_hardened),("iip",collect_iip_hardened)):
        try:
            result=func(); status.setdefault("stages",{})[name]=result; print(name,"OK",json.dumps(result,ensure_ascii=False))
        except Exception as exc:
            status.setdefault("stages",{})[name]={"status":"error","error":f"{type(exc).__name__}: {exc}"}; print(name,"ERROR",status["stages"][name]["error"])
    status["updated_at"]=now_iso()
    status["status"]="success" if all(v.get("status") in {"success","delegated"} for v in status.get("stages",{}).values()) else "partial"
    save_json("data/economy/historical-backfill-status.json",status)
    save_json("data/economy/historical-coverage.json",build_coverage(status.get("stages",{})))
    print(json.dumps(status,ensure_ascii=False,indent=2))

if __name__=="__main__": main()
