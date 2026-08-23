#!/usr/bin/env python3
"""Robust entry point for the Energy Agency long-run weekly fuel workbook.

The official site currently names the long-run filling-station weekly workbook
`YYMMDDs5.xlsx`; the current-week detail workbook is `YYMMDD.xlsx`.
For XLSX, rows are streamed into memory once before the layout parser runs.
This avoids repeated random `ws.cell()` access on openpyxl read-only sheets,
which is prohibitively slow on the 1990-present workbook.
"""
from __future__ import annotations

import io, json, re, traceback
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from update_japan_fuel_prices import RESULTS, UA, write_status, parse_sheet, parse_xls

ROOT = Path(__file__).resolve().parents[1]


def find_long_run_workbook(content: bytes):
    soup = BeautifulSoup(content, "html.parser")
    excel = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if re.search(r"\.xlsx?(?:\?|$)", href, re.I):
            excel.append(urljoin(RESULTS, href))
    s5 = [u for u in excel if re.search(r"s5\.xlsx?(?:\?|$)", u, re.I)]
    if s5:
        return s5[0], excel
    raise RuntimeError(f"official long-run s5 weekly workbook not found; excel_links={excel[:20]}")


def parse_xlsx_fast(content: bytes):
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parsed, details = {}, []
    for ws in wb.worksheets:
        matrix = [tuple(row) for row in ws.iter_rows(values_only=True)]
        if not matrix:
            continue
        nrows = len(matrix)
        ncols = max((len(row) for row in matrix), default=0)
        def getter(r, c):
            if r < 0 or r >= nrows or c < 0 or c >= len(matrix[r]):
                return None
            return matrix[r][c]
        result = parse_sheet(ws.title, nrows, ncols, getter)
        if result:
            mid, rows, layout = result
            if len(rows) > len(parsed.get(mid, [])):
                parsed[mid] = rows
            details.append({"sheet": ws.title, "metric_id": mid, "layout": layout, "rows": len(rows)})
    wb.close()
    return parsed, details


def main():
    weekly = None
    try:
        write_status("running", "fetch_index_v2")
        page = requests.get(RESULTS, headers=UA, timeout=30); page.raise_for_status()
        weekly, candidates = find_long_run_workbook(page.content)
        write_status("running", "download_workbook_v2", weekly_workbook_url=weekly, link_candidates=candidates[:20])
        book = requests.get(weekly, headers=UA, timeout=90); book.raise_for_status(); content = book.content
        if content.startswith(b"PK"):
            workbook_format, (parsed, details) = "xlsx", parse_xlsx_fast(content)
        elif content.startswith(bytes.fromhex("D0CF11E0A1B11AE1")):
            workbook_format, (parsed, details) = "xls", parse_xls(content)
        else:
            raise RuntimeError(f"unsupported workbook signature: {content[:16].hex()}")
        required = ("regular_gasoline_national", "diesel_national")
        if not all(parsed.get(mid) for mid in required):
            raise RuntimeError(f"required national series missing; found={list(parsed)} details={details}")

        path = ROOT / "data/economy/fuel-prices.json"
        data = json.loads(path.read_text(encoding="utf-8")); by_id = {s["metric_id"]:s for s in data.get("series",[])}
        for mid, rows in parsed.items():
            if mid in by_id and rows: by_id[mid]["observations"] = rows
        data["collection_status"] = "official_weekly_history_ingested"
        data["source"]["weekly_workbook_url"] = weekly
        data["last_collected_at"] = datetime.now(timezone.utc).isoformat()
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        counts = {mid:len(rows) for mid,rows in parsed.items()}
        write_status("success","success",weekly_workbook_url=weekly,workbook_format=workbook_format,
                     parsed_sheets=details,row_counts=counts)
        print(counts)
    except Exception as exc:
        write_status("failure","collector_v2",weekly_workbook_url=weekly,error=str(exc),diagnostic_tail=traceback.format_exc()[-6000:])
        raise

if __name__ == "__main__": main()
