#!/usr/bin/env python3
"""Temporary diagnostic for the MLIT 21-company warehouse workbook layout."""
from io import BytesIO
from urllib.parse import urljoin
import re
import requests
from bs4 import BeautifulSoup
import xlrd
from openpyxl import load_workbook

PAGE="https://www.mlit.go.jp/seisakutokatsu/freight/seisakutokatsu_freight_mn2_000009.html"
UA={"User-Agent":"Mozilla/5.0 LBI-Warehouse-Diagnostic/1.0"}
html=requests.get(PAGE,headers=UA,timeout=60); html.raise_for_status()
soup=BeautifulSoup(html.text,"html.parser")
links=[]
for a in soup.find_all("a",href=True):
    href=urljoin(PAGE,a["href"])
    if re.search(r"\.xls[x]?(?:$|\?)",href,re.I): links.append((a.get_text(" ",strip=True),href))
print("excel links",len(links))
label,url=links[0]
print("selected",label,url)
blob=requests.get(url,headers=UA,timeout=60); blob.raise_for_status()
keywords=("入庫","出庫","保管残高","普通倉庫","合計","回転率","倉庫利用")

def compact(values):
    return " | ".join(str(v).strip() for v in values if v not in (None,""))

if url.lower().split("?")[0].endswith(".xlsx"):
    book=load_workbook(BytesIO(blob.content),data_only=True,read_only=True)
    print("sheets",book.sheetnames)
    for sname in book.sheetnames:
        ws=book[sname]
        hits=[]
        for r_idx,row in enumerate(ws.iter_rows(values_only=True),1):
            joined=compact(row)
            if any(k in joined for k in keywords): hits.append((r_idx,joined[:1600]))
        if hits:
            print("SHEET",sname,"rows",ws.max_row,"cols",ws.max_column)
            for r,text in hits[:100]: print(f"R{r}: {text}")
    ws=book["推移表"]
    print("TREND TABLE FIRST 45 ROWS A:T")
    for r in range(1,min(ws.max_row,45)+1):
        print(f"T{r}: {compact([ws.cell(r,c).value for c in range(1,21)])}")
    print("TREND TABLE NONEMPTY COLUMN HEADERS ROWS 1-4 ACROSS ALL COLS")
    for c in range(1,ws.max_column+1):
        vals=[ws.cell(r,c).value for r in range(1,5)]
        if any(v not in (None,"") for v in vals): print(f"C{c}: {compact(vals)}")
    print("TREND TABLE LAST 12 NONEMPTY MONTHLY ROWS A:T")
    nonempty=[]
    for r in range(31,ws.max_row+1):
        vals=[ws.cell(r,c).value for c in range(1,21)]
        if any(v not in (None,"") for v in vals): nonempty.append((r,vals))
    for r,vals in nonempty[-12:]: print(f"L{r}: {compact(vals)}")
else:
    book=xlrd.open_workbook(file_contents=blob.content)
    print("sheets",book.sheet_names())
    for sname in book.sheet_names():
        ws=book.sheet_by_name(sname)
        hits=[]
        for r in range(ws.nrows):
            vals=[ws.cell_value(r,c) for c in range(ws.ncols)]
            joined=compact(vals)
            if any(k in joined for k in keywords): hits.append((r+1,joined[:1600]))
        if hits:
            print("SHEET",sname,"rows",ws.nrows,"cols",ws.ncols)
            for r,text in hits[:100]: print(f"R{r}: {text}")
