#!/usr/bin/env python3
import io, requests
from openpyxl import load_workbook
URL='https://www.e-stat.go.jp/stat-search/file-download?fileKind=0&statInfId=000040172363'
r=requests.get(URL,headers={'User-Agent':'Mozilla/5.0 LBI-Historical-Backfill/1.0','Referer':'https://www.e-stat.go.jp/'},timeout=90);r.raise_for_status()
wb=load_workbook(io.BytesIO(r.content),read_only=True,data_only=True)
for ws in wb.worksheets:
    print('SHEET',ws.title,'rows',ws.max_row,'cols',ws.max_column)
    for i,row in enumerate(ws.iter_rows(values_only=True)):
        if i>=35:break
        print(i+1,repr(tuple(row[:18])))
wb.close()
